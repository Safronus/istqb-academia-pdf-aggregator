from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from PySide6.QtCore import QSortFilterProxyModel, Qt, QItemSelectionModel, QModelIndex
from PySide6.QtGui import QAction, QDesktopServices
from PySide6.QtWidgets import (
    QWidget, QMainWindow, QTabWidget, QVBoxLayout, QHBoxLayout,
    QTableView, QLineEdit, QPushButton, QLabel, QComboBox, QFormLayout,
    QTreeView, QFileSystemModel, QSplitter, QMessageBox
)

from .pdf_scanner import PdfScanner, PdfRecord
from .istqb_boards import KNOWN_BOARDS














from PySide6.QtCore import Qt
from PySide6.QtWidgets import QStyledItemDelegate, QStyleOptionViewItem, QApplication, QStyle

class BoardHidingDelegate(QStyledItemDelegate):
    """
    Vykreslovací delegát pro sloupec 'Board':
    - Pro PRVNÍ výskyt hodnoty v aktuálním pořadí tabulky vykreslí text normálně.
    - Pro DALŠÍ výskyty stejné hodnoty vykreslí prázdný text (jen vizuálně).
    - Nezasahuje do modelu, dat ani exportu (data(DisplayRole) zůstávají zachována).
    """
    def paint(self, painter, option, index):
        if index.column() == 0:  # sloupec Board
            current = index.data(Qt.DisplayRole)
            hide = False
            # Zjisti, zda se stejná hodnota vyskytla již výše (v proxy pořadí)
            # Pozn.: lineární průchod jen přes viditelný model; výkonově OK pro běžné tabulky.
            try:
                model = index.model()
                for r in range(0, index.row()):
                    other = model.index(r, index.column()).data(Qt.DisplayRole)
                    if other == current and current not in (None, ""):
                        hide = True
                        break
            except Exception:
                hide = False

            if hide:
                opt = QStyleOptionViewItem(option)
                self.initStyleOption(opt, index)
                opt.text = ""  # vykreslit prázdný text, zbytek styly necháme
                style = option.widget.style() if option.widget else QApplication.style()
                style.drawControl(QStyle.CE_ItemViewItem, opt, painter, option.widget)
                return

        # Defaultní vykreslení pro ostatní sloupce nebo první výskyt Board
        super().paint(painter, option, index)
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        

class RecordsModel(QSortFilterProxyModel):
    def __init__(self, headers: List[str], parent=None):
        super().__init__(parent)
        self.headers = headers
        self.search = ""
        self.board_filter = "All"

    def set_search(self, text: str) -> None:
        self.search = text.lower().strip()
        self.invalidateFilter()

    def set_board(self, board: str) -> None:
        self.board_filter = board
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        model = self.sourceModel()
        if model is None:
            return True
        # Board filter is based on column 0 (Board)
        idx_board = model.index(source_row, 0, source_parent)
        board_val = (model.data(idx_board, Qt.DisplayRole) or "").strip()
        if self.board_filter != "All" and board_val != self.board_filter:
            return False

        if not self.search:
            return True

        # Search across all columns
        for c in range(model.columnCount()):
            idx = model.index(source_row, c, source_parent)
            val = (model.data(idx, Qt.DisplayRole) or "").lower()
            if self.search in val:
                return True
        return False

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:
        """Řazení: Board → Application Type → Candidate Name."""
        model = self.sourceModel()
        if model is None:
            return super().lessThan(left, right)
    
        def data(row: int, col: int) -> str:
            idx = model.index(row, col)
            return (model.data(idx, Qt.DisplayRole) or "").strip()
    
        BOARD = 0
        APP   = 1
        CAND  = 3
    
        a_board = data(left.row(), BOARD).lower()
        b_board = data(right.row(), BOARD).lower()
    
        order = {"new application": 0, "additional recognition": 1}
        a_app = order.get(data(left.row(), APP).lower(), 2)
        b_app = order.get(data(right.row(), APP).lower(), 2)
    
        a_cand = data(left.row(), CAND).lower()
        b_cand = data(right.row(), CAND).lower()
    
        return (a_board, a_app, a_cand) < (b_board, b_app, b_cand)

class MainWindow(QMainWindow):
    def __init__(self, pdf_root: Path):
        super().__init__()
        self.setWindowTitle("ISTQB Academia PDF Aggregator")
        self.resize(1200, 800)
        self.pdf_root = pdf_root
    
        # Kořen "Sorted PDFs"
        self.sorted_root = (Path.cwd() / "Sorted PDFs").resolve()
    
        # DB pro Sorted PDFs
        from app.sorted_db import SortedDb
        self.sorted_db = SortedDb(self.sorted_root)
        self.sorted_db.load()
    
        self.records: List[PdfRecord] = []
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
    
        self.overview_tab = QWidget()
        self.browser_tab = QWidget()
        self.sorted_tab = QWidget()
        self.contacts_tab = QWidget()
        self.recognized_tab = QWidget()
    
        self.tabs.addTab(self.overview_tab, "Overview")
        self.tabs.addTab(self.browser_tab, "PDF Browser")
        self.tabs.addTab(self.sorted_tab, "Sorted PDFs")
        self.tabs.addTab(self.contacts_tab, "Board Contacts")
        self.tabs.addTab(self.recognized_tab, "Recognized People List")
    
        self._build_menu()
        self._build_overview_tab()
        self._build_browser_tab()
        self._build_sorted_tab()
        self._build_contacts_tab()
        self._build_recognized_tab()
    
        self.rescan()
        self.rescan_sorted()
        self._init_fs_watcher()
        
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self._apply_global_sizing_once)

    def _build_recognized_tab(self) -> None:
        """
        Recognized People List:
          - Tabulka: Board, Full Name, Email, Address, Recognition Date, Valid Until, Badge Types, Badge Link
          - Filtrace: fulltext + checkboxy (Valid / Near expiry / Expired)
          - Akce: Add…, Edit…, Delete, Reload, Save
          - JSON perzistence (recognized_people.json)
          - Fitting sloupců + barevné zvýraznění řádků při přepnutí do záložky
        """
        from PySide6.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QWidget, QTableView, QPushButton, QHeaderView,
            QLineEdit, QCheckBox, QLabel
        )
        from PySide6.QtGui import QStandardItemModel
        from PySide6.QtCore import Qt, QTimer, QSortFilterProxyModel
        from datetime import datetime, date
    
        layout = QVBoxLayout(self.recognized_tab)
    
        # Ovládací lišta
        bar = QHBoxLayout()
        self.btn_rec_add = QPushButton("Add person…")
        self.btn_rec_edit = QPushButton("Edit…")
        self.btn_rec_del  = QPushButton("Delete")
        self.btn_rec_reload = QPushButton("Reload")
        self.btn_rec_save = QPushButton("Save")
        bar.addWidget(self.btn_rec_add)
        bar.addWidget(self.btn_rec_edit)
        bar.addWidget(self.btn_rec_del)
        bar.addStretch(1)
        bar.addWidget(self.btn_rec_reload)
        bar.addWidget(self.btn_rec_save)
        layout.addLayout(bar)
    
        # FILTRAČNÍ ŘÁDEK (fulltext + 3 checkboxy)
        filt = QHBoxLayout()
        self.rec_search = QLineEdit(self.recognized_tab)
        self.rec_search.setPlaceholderText("Search…")
        self.rec_chk_valid   = QCheckBox("Valid")
        self.rec_chk_warning = QCheckBox("Near expiry")
        self.rec_chk_expired = QCheckBox("Expired")
        self.rec_chk_valid.setChecked(True)
        self.rec_chk_warning.setChecked(True)
        self.rec_chk_expired.setChecked(True)
        filt.addWidget(QLabel("Filter:"))
        filt.addWidget(self.rec_search, 1)
        filt.addSpacing(12)
        filt.addWidget(self.rec_chk_valid)
        filt.addWidget(self.rec_chk_warning)
        filt.addWidget(self.rec_chk_expired)
        layout.addLayout(filt)
    
        # Tabulka + MODEL
        self.tbl_recognized = QTableView(self.recognized_tab)
        self.tbl_recognized.setSelectionBehavior(QTableView.SelectRows)
        self.tbl_recognized.setSelectionMode(QTableView.ExtendedSelection)
        self.tbl_recognized.setSortingEnabled(True)
    
        # Pořadí sloupců (0.10d)
        self._recognized_headers = [
            "Board", "Full Name", "Email", "Address",
            "Recognition Date", "Valid Until", "Badge Types", "Badge Link"
        ]
        self._recognized_model = QStandardItemModel(0, len(self._recognized_headers), self)
        self._recognized_model.setHorizontalHeaderLabels(self._recognized_headers)
    
        # PROXY MODEL pro filtraci
        class _RecognizedProxy(QSortFilterProxyModel):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.search = ""
                self.show_valid = True
                self.show_warn = True
                self.show_expired = True
    
            def _status_for_row(self, row: int) -> str:
                """Vrátí 'valid' / 'warn' / 'expired' dle 'Valid Until' (col 5)."""
                try:
                    idx = self.sourceModel().index(row, 5)
                    s = self.sourceModel().data(idx)
                    if not s:
                        return "valid"
                    vuntil = datetime.strptime(str(s), "%Y-%m-%d").date()
                    days_left = (vuntil - date.today()).days
                    if days_left > 30:
                        return "valid"
                    elif days_left >= 0:
                        return "warn"
                    else:
                        return "expired"
                except Exception:
                    return "valid"
    
            def filterAcceptsRow(self, source_row: int, source_parent) -> bool:
                # 1) Validitní checkboxy
                st = self._status_for_row(source_row)
                if st == "valid" and not self.show_valid:
                    return False
                if st == "warn" and not self.show_warn:
                    return False
                if st == "expired" and not self.show_expired:
                    return False
                # 2) Fulltext přes všechny sloupce
                if self.search:
                    s = self.search.lower()
                    cols = self.sourceModel().columnCount()
                    found = False
                    for c in range(cols):
                        idx = self.sourceModel().index(source_row, c)
                        val = self.sourceModel().data(idx)
                        if val and s in str(val).lower():
                            found = True
                            break
                    if not found:
                        return False
                return True
    
        self._rec_proxy = _RecognizedProxy(self)
        self._rec_proxy.setSourceModel(self._recognized_model)
        self.tbl_recognized.setModel(self._rec_proxy)
    
        # Fit sloupců
        hdr = self.tbl_recognized.horizontalHeader()
        hdr.setStretchLastSection(True)
        try:
            hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        except Exception:
            try:
                hdr.setResizeMode(0, QHeaderView.ResizeToContents)
            except Exception:
                pass
    
        layout.addWidget(self.tbl_recognized, 1)
    
        # Naplň data
        self._recognized_rebuild_model()
    
        # Akce
        self.btn_rec_add.clicked.connect(self._recognized_add)
        self.btn_rec_edit.clicked.connect(self._recognized_edit)
        self.btn_rec_del.clicked.connect(self._recognized_delete)
        self.btn_rec_reload.clicked.connect(self._recognized_rebuild_model)
        self.btn_rec_save.clicked.connect(lambda: self._save_recognized_json(self._recognized_collect_data()))
    
        # Filtrační signály
        self.rec_search.textChanged.connect(lambda _t: self._recognized_update_filter())
        self.rec_chk_valid.toggled.connect(lambda _b: self._recognized_update_filter())
        self.rec_chk_warning.toggled.connect(lambda _b: self._recognized_update_filter())
        self.rec_chk_expired.toggled.connect(lambda _b: self._recognized_update_filter())
    
        # Po vykreslení dofituj a aplikuj barvy
        QTimer.singleShot(0, self._recognized_fit_columns)
        QTimer.singleShot(0, self._recognized_apply_row_colors)
    
        # Spusť zvýraznění při přepnutí do záložky (napoj jednorázově)
        if hasattr(self, "tabs") and not getattr(self, "_rec_tab_hooked", False):
            try:
                self.tabs.currentChanged.connect(self._recognized_on_tab_changed)
                self._rec_tab_hooked = True
            except Exception:
                pass
            
    def _recognized_update_filter(self) -> None:
        """Aplikuje hodnoty z fulltextu a checkboxů do proxy a invaliduje filtr."""
        try:
            proxy = getattr(self, "_rec_proxy", None)
            if not proxy:
                return
            proxy.search = (self.rec_search.text() or "").strip()
            proxy.show_valid = bool(self.rec_chk_valid.isChecked())
            proxy.show_warn = bool(self.rec_chk_warning.isChecked())
            proxy.show_expired = bool(self.rec_chk_expired.isChecked())
            proxy.invalidateFilter()
        except Exception:
            pass
            
    def _recognized_on_tab_changed(self, index: int) -> None:
        """
        Při přepnutí na Recognized tab obnov barvy (a můžeš i dofit sloupce).
        """
        try:
            w = self.tabs.widget(index)
            if w is self.recognized_tab:
                self._recognized_apply_row_colors()
                self._recognized_fit_columns()
        except Exception:
            pass

    def _recognized_json_path(self):
        """Cesta k JSONu s recognized osobami (repo root)."""
        from pathlib import Path
        return Path(__file__).resolve().parents[2] / "recognized_people.json"

    def _load_recognized_json(self) -> list[dict]:
        """Načti JSON (list dictů). Neexistuje-li, vrať []."""
        import json
        p = self._recognized_json_path()
        try:
            if p.exists():
                with p.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    if isinstance(data, list):
                        return data
        except Exception:
            pass
        return []
    
    def _save_recognized_json(self, data: list[dict]) -> None:
        """Ulož JSON s recognized osobami."""
        import json
        p = self._recognized_json_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            with p.open("w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False, indent=2)
            try:
                self.statusBar().showMessage("Recognized people saved.")
            except Exception:
                pass
        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Save recognized", f"Failed to save recognized_people.json:\n{e}")    
    
    def _recognized_rebuild_model(self) -> None:
        """
        Načti JSON a naplň model.
        - Pokud záznam obsahuje oba badge (academia & certified), zobraz ho jako DVA řádky (sdruženě).
        - 'Valid Until' = 'Recognition Date' + 365 dní (přepočet na load).
        - POŘADÍ SLOUPCŮ: Board, Full Name, Email, Address, Recognition Date, Valid Until, Badge Types, Badge Link
        """
        from PySide6.QtGui import QStandardItem
        from PySide6.QtCore import Qt
        from datetime import datetime, timedelta
    
        data = self._load_recognized_json()
        self._recognized_model.setRowCount(0)
    
        def _valid_until(date_str: str) -> str:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d").date()
                return (dt + timedelta(days=365)).isoformat()
            except Exception:
                return ""
    
        def _append_row(board, full, mail, addr, rdate, acad, cert, blink):
            badges = "; ".join([s for s, b in (("Academia", acad), ("Certified", cert)) if b])
            vuntil = _valid_until(rdate)
            # nový pořadník
            vals = [board, full, mail, addr, rdate, vuntil, badges, blink]
            items = [QStandardItem(str(v)) for v in vals]
            for it in items:
                it.setEditable(False)
            self._recognized_model.appendRow(items)
    
        for rec in data:
            board = rec.get("board", "") or ""
            full  = rec.get("full_name", "") or ""
            mail  = rec.get("email", "") or ""
            addr  = rec.get("address", "") or ""
            rdate = rec.get("recognition_date", "") or ""
            acad  = bool(rec.get("academia", False))
            cert  = bool(rec.get("certified", False))
            blink = rec.get("badge_link", "") or ""
    
            if acad and cert:
                _append_row(board, full, mail, addr, rdate, True,  False, blink)
                _append_row(board, full, mail, addr, rdate, False, True,  blink)
            else:
                _append_row(board, full, mail, addr, rdate, acad, cert, blink)
    
        self._recognized_fit_columns()
    
    def _recognized_collect_data(self) -> list[dict]:
        """Převeď model -> JSON list."""
        from PySide6.QtCore import Qt
        out = []
        for r in range(self._recognized_model.rowCount()):
            board = self._recognized_model.index(r, 0).data(Qt.DisplayRole) or ""
            full  = self._recognized_model.index(r, 1).data(Qt.DisplayRole) or ""
            mail  = self._recognized_model.index(r, 2).data(Qt.DisplayRole) or ""
            addr  = self._recognized_model.index(r, 3).data(Qt.DisplayRole) or ""
            rdate = self._recognized_model.index(r, 4).data(Qt.DisplayRole) or ""
            badges= (self._recognized_model.index(r, 5).data(Qt.DisplayRole) or "").lower()
            blink = self._recognized_model.index(r, 6).data(Qt.DisplayRole) or ""
            vuntil= self._recognized_model.index(r, 7).data(Qt.DisplayRole) or ""
    
            acad = "academia" in badges
            cert = "certified" in badges
    
            out.append({
                "board": board,
                "full_name": full,
                "email": mail,
                "address": addr,
                "recognition_date": rdate,
                "academia": acad,
                "certified": cert,
                "badge_link": blink,
                "valid_until": vuntil,
            })
        return out    
    
    def _recognized_candidates_from_sorted(self) -> list[dict]:
        """
        Najdi kandidáty pro Recognized Add dialog.
        Preferujeme 'Sorted DB' v paměti; pokud není, zkusíme běžné JSON cesty.
        Jako poslední fallback odvodíme kandidáty z Overview tabulky.
    
        Vrací list dictů:
          { "board": str, "full_name": str, "email": str, "address": str,
            "academia": bool, "certified": bool }
        """
        from pathlib import Path
        candidates: list[dict] = []
        seen: set[tuple] = set()
    
        def _add(board, full, mail, addr, acad, cert):
            key = (str(board or ""), str(full or ""), str(mail or ""), str(addr or ""), bool(acad), bool(cert))
            if key in seen:
                return
            seen.add(key)
            candidates.append({
                "board": str(board or ""),
                "full_name": str(full or ""),
                "email": str(mail or ""),
                "address": str(addr or ""),
                "academia": bool(acad),
                "certified": bool(cert),
            })
    
        # 1) In-memory zdroje – různé názvy používané v projektu
        for attr in ("sorted_db", "_sorted_db", "sorted_records", "_sorted_records", "sorted_data", "_sorted_data"):
            container = getattr(self, attr, None)
            if not container:
                continue
    
            def _iter(container):
                if isinstance(container, dict):
                    for v in container.values():
                        yield v
                elif isinstance(container, list):
                    for v in container:
                        yield v
    
            for rec in _iter(container):
                # rec může být objekt i dict
                try:
                    get = (lambda k, default="": getattr(rec, k, getattr(rec, k.replace(" ", "_"), None))
                           if hasattr(rec, k) or hasattr(rec, k.replace(" ", "_"))
                           else rec.get(k, rec.get(k.replace(" ", "_"), default)))
                except Exception:
                    # rec je nejspíš dict
                    def get(k, default=""):
                        return rec.get(k, rec.get(k.replace(" ", "_"), default))
    
                board = get("board", "")
                full  = get("contact_full_name", "") or get("full_name", "")
                mail  = get("contact_email", "") or get("email", "")
                addr  = get("contact_postal_address", "") or get("address", "")
    
                acad_raw = get("recognition_academia", "")
                cert_raw = get("recognition_certified", "")
                # Interpretuj jakýkoli neprázdný string/true hodnotu jako True
                acad = bool(acad_raw) and str(acad_raw).strip().lower() not in ("false", "0", "no", "none")
                cert = bool(cert_raw) and str(cert_raw).strip().lower() not in ("false", "0", "no", "none")
    
                if full or mail:
                    _add(board, full, mail, addr, acad, cert)
    
            if candidates:
                return candidates  # máme výsledky, dál nehledáme
    
        # 2) JSON soubory – pár běžných cest/jmen (pokud existují)
        try_paths = []
        try:
            repo_root = Path(__file__).resolve().parents[2]
            try_paths.extend([
                repo_root / "sorted_db.json",
                repo_root / "sorted_records.json",
                repo_root / "data" / "sorted_db.json",
                repo_root / "Sorted PDFs" / "sorted_db.json",
            ])
        except Exception:
            pass
    
        import json
        for p in try_paths:
            try:
                if p.exists():
                    with p.open("r", encoding="utf-8") as fh:
                        data = json.load(fh)
                    # projdi stejným způsobem jako in-memory
                    def _iter2(container):
                        if isinstance(container, dict):
                            for v in container.values():
                                yield v
                        elif isinstance(container, list):
                            for v in container:
                                yield v
                    for rec in _iter2(data):
                        if isinstance(rec, dict):
                            board = rec.get("board", "")
                            full  = rec.get("contact_full_name", "") or rec.get("full_name", "")
                            mail  = rec.get("contact_email", "") or rec.get("email", "")
                            addr  = rec.get("contact_postal_address", "") or rec.get("address", "")
                            acad_raw = rec.get("recognition_academia", "")
                            cert_raw = rec.get("recognition_certified", "")
                            acad = bool(acad_raw) and str(acad_raw).strip().lower() not in ("false", "0", "no", "none")
                            cert = bool(cert_raw) and str(cert_raw).strip().lower() not in ("false", "0", "no", "none")
                            if full or mail:
                                _add(board, full, mail, addr, acad, cert)
                    if candidates:
                        return candidates
            except Exception:
                continue
    
        # 3) Fallback: odvoď z Overview tabulky (_source_model) — pokud existuje
        for d in self._overview_iter_records_as_dicts():
            board = d.get("board", "")
            full  = d.get("contact_full_name", "") or d.get("full_name", "")
            mail  = d.get("contact_email", "") or d.get("email", "")
            addr  = d.get("contact_postal_address", "") or d.get("address", "")
            acad_raw = d.get("recognition_academia", "")
            cert_raw = d.get("recognition_certified", "")
            acad = bool(acad_raw) and str(acad_raw).strip().lower() not in ("false", "0", "no", "none")
            cert = bool(cert_raw) and str(cert_raw).strip().lower() not in ("false", "0", "no", "none")
            if full or mail:
                _add(board, full, mail, addr, acad, cert)
    
        return candidates
    
    def _overview_iter_records_as_dicts(self) -> list[dict]:
        """
        Přečti data z Overview (pokud existuje _source_model) a vrať list dictů
        se standardizovanými klíči: board, contact_full_name, contact_email,
        contact_postal_address, recognition_academia, recognition_certified.
        Pokud model neexistuje, vrať [].
        """
        from PySide6.QtCore import Qt
        out: list[dict] = []
        src = getattr(self, "_source_model", None)
        if not src:
            return out
    
        # mapuj indexy sloupců podle _headers
        idx = {}
        headers = getattr(self, "_headers", [])
        def _find(col_label_suffix: str) -> int | None:
            # hledáme poslední řádek popisku po \n
            for i, h in enumerate(headers):
                tail = h.split("\n")[-1].strip() if "\n" in h else h.strip()
                if tail.lower() == col_label_suffix.lower():
                    return i
            return None
    
        col_board = _find("Board")
        col_full  = _find("Full Name")
        col_mail  = _find("Email Address")
        col_addr  = _find("Postal Address")
        col_acad  = _find("Academia Recognition")
        col_cert  = _find("Certified Recognition")
    
        rows = src.rowCount()
        for r in range(rows):
            def _val(c):
                if c is None: return ""
                return src.index(r, c).data(Qt.DisplayRole) or ""
            out.append({
                "board": _val(col_board),
                "contact_full_name": _val(col_full),
                "contact_email": _val(col_mail),
                "contact_postal_address": _val(col_addr),
                "recognition_academia": _val(col_acad),
                "recognition_certified": _val(col_cert),
            })
        return out
    
    def _recognized_open_add_dialog(self, initial: dict | None = None) -> dict | None:
        """
        Add/Edit dialog pro recognized osobu.
        - ADD: volba zdroje (From Sorted PDFs / Manual), badge typy editovatelné.
        - EDIT: zdroj i badge typy jsou irelevantní → skryté/disable; typy zůstávají dle řádku.
        Vrací dict se stejnými klíči jako JSON nebo None při Cancel.
        """
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, QLineEdit, QPlainTextEdit,
            QDialogButtonBox, QRadioButton, QComboBox, QDateEdit, QCheckBox, QMessageBox
        )
        from PySide6.QtCore import Qt, QDate
        from datetime import timedelta, datetime as _dt
    
        is_edit = initial is not None
    
        dlg = QDialog(self)
        dlg.setWindowTitle("Edit recognized person" if is_edit else "Add recognized person")
        lay = QVBoxLayout(dlg)
    
        # Zdroj dat (Sorted vs Manual) – v EDIT režimu skryjeme
        src_row = QHBoxLayout()
        rb_sorted = QRadioButton("From Sorted PDFs")
        rb_manual = QRadioButton("Manual entry")
        if not is_edit:
            src_row.addWidget(rb_sorted); src_row.addWidget(rb_manual); src_row.addStretch(1)
            lay.addLayout(src_row)
    
        # Kandidáti ze Sorted DB – v EDIT režimu neukazujeme
        cb_sorted = QComboBox()
        lbl_no_sorted = QLabel("No candidates found in Sorted PDFs database.")
        lbl_no_sorted.setStyleSheet("color: #999;")
        if not is_edit:
            candidates = self._recognized_candidates_from_sorted()
            for c in candidates:
                disp = f"{c.get('full_name','')} — {c.get('email','')} — [{c.get('board','')}]"
                cb_sorted.addItem(disp, c)
            lay.addWidget(cb_sorted)
            lay.addWidget(lbl_no_sorted)
            # výchozí stav
            if candidates:
                rb_sorted.setChecked(True)
            else:
                rb_manual.setChecked(True)
    
        # Form
        form = QFormLayout()
        ed_board = QLineEdit()
        ed_full  = QLineEdit()
        ed_mail  = QLineEdit()
        ed_addr  = QPlainTextEdit(); ed_addr.setMinimumHeight(64)
    
        ed_date  = QDateEdit(); ed_date.setCalendarPopup(True); ed_date.setDate(QDate.currentDate())
        chk_acad = QCheckBox("Academia Recognition")
        chk_cert = QCheckBox("Certified Recognition")
        ed_link  = QLineEdit()
    
        form.addRow("Board:", ed_board)
        form.addRow("Full Name:", ed_full)
        form.addRow("Email:", ed_mail)
        form.addRow("Address:", ed_addr)
        form.addRow("Recognition Date:", ed_date)
        # Badge types – v EDIT režimu jen zobrazíme/disable, v ADD editovatelné
        if not is_edit:
            form.addRow("Badge Types:", QLabel(""))
            form.addRow("", chk_acad)
            form.addRow("", chk_cert)
        else:
            # jen informativní řádek s textem badge, bez checků
            self._lbl_badges_info = QLabel("(badge type unchanged in Edit)")
            form.addRow("Badge Types:", self._lbl_badges_info)
            chk_acad.setVisible(False)
            chk_cert.setVisible(False)
    
        form.addRow("Badge Link:", ed_link)
        lay.addLayout(form)
    
        # Předvyplnění z initial (Edit)
        if is_edit:
            ed_board.setText(initial.get("board", ""))
            ed_full.setText(initial.get("full_name", ""))
            ed_mail.setText(initial.get("email", ""))
            ed_addr.setPlainText(initial.get("address", ""))
            try:
                y, m, d = map(int, (initial.get("recognition_date","") or "2000-01-01").split("-"))
                ed_date.setDate(QDate(y, m, d))
            except Exception:
                pass
            # Badge typy ponecháváme; jen je držíme v initial
            chk_acad.setChecked(bool(initial.get("academia", False)))
            chk_cert.setChecked(bool(initial.get("certified", False)))
            ed_link.setText(initial.get("badge_link", ""))
        else:
            # ADD: předvyplnění po výběru kandidáta ze Sorted
            def _apply_candidate():
                data = cb_sorted.currentData()
                if not data:
                    return
                ed_board.setText(data.get("board",""))
                ed_full.setText(data.get("full_name",""))
                ed_mail.setText(data.get("email",""))
                ed_addr.setPlainText(data.get("address",""))
                chk_acad.setChecked(bool(data.get("academia", False)))
                chk_cert.setChecked(bool(data.get("certified", False)))
            cb_sorted.currentIndexChanged.connect(_apply_candidate)
            # přepínání zdroje
            def _toggle_src():
                use_sorted = rb_sorted.isChecked() if not is_edit else False
                has_data = (cb_sorted.count() > 0) if not is_edit else False
                if not is_edit:
                    cb_sorted.setEnabled(use_sorted and has_data)
                    lbl_no_sorted.setVisible(use_sorted and not has_data)
            if not is_edit:
                rb_sorted.toggled.connect(_toggle_src)
                rb_manual.toggled.connect(_toggle_src)
                # init
                if cb_sorted.count() > 0:
                    rb_sorted.setChecked(True)
                    _apply_candidate()
                _toggle_src()
    
        # Tlačítka
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dlg)
        lay.addWidget(btns)
    
        result: dict | None = None
    
        def _accept():
            # Collect
            board = ed_board.text().strip()
            full  = ed_full.text().strip()
            mail  = ed_mail.text().strip()
            addr  = ed_addr.toPlainText().strip()
            date_str = ed_date.date().toString("yyyy-MM-dd")
            # v EDIT režimu použij původní typy, v ADD podle checkboxů
            if is_edit:
                acad = bool(initial.get("academia", False))
                cert = bool(initial.get("certified", False))
            else:
                acad = chk_acad.isChecked()
                cert = chk_cert.isChecked()
            link = ed_link.text().strip()
    
            # Validate
            if not full:
                QMessageBox.information(dlg, "Validation", "Full Name is required.")
                return
            if not (acad or cert):
                QMessageBox.information(dlg, "Validation", "Select at least one badge type.")
                return
            if not date_str:
                QMessageBox.information(dlg, "Validation", "Recognition Date is required.")
                return
            if not link:
                QMessageBox.information(dlg, "Validation", "Badge Link is required.")
                return
    
            # Valid Until (+365 dní) – počítáme až při ukládání do modelu
            nonlocal result
            result = {
                "board": board,
                "full_name": full,
                "email": mail,
                "address": addr,
                "recognition_date": date_str,
                "academia": bool(acad),
                "certified": bool(cert),
                "badge_link": link,
            }
            dlg.accept()
    
        btns.accepted.connect(_accept)
        btns.rejected.connect(dlg.reject)
    
        dlg.resize(640, 520)
        return result if dlg.exec_() == QDialog.Accepted else None
    
    def _recognized_add(self) -> None:
        """
        Přidání osoby. Při volbě obou badge vloží DVA řádky (po kontrole duplicit per-badge).
        Duplikát: (full_name + badge_link) NEBO (full_name + recognition_date + badge-typ).
        Nový pořádek sloupců: Board, Full Name, Email, Address, Recognition Date, Valid Until, Badge Types, Badge Link.
        """
        from PySide6.QtWidgets import QMessageBox
        from PySide6.QtGui import QStandardItem
        from PySide6.QtCore import Qt
        from datetime import datetime, timedelta
    
        new = self._recognized_open_add_dialog(None)
        if not new:
            return
    
        def _badge_rows(n: dict) -> list[tuple[bool, bool]]:
            a = bool(n.get("academia"))
            c = bool(n.get("certified"))
            if a and c:
                return [(True, False), (False, True)]
            return [(a, c)]
    
        def _dup_exists(full_l, link_l, dstr, acad, cert) -> bool:
            bkey = f"{int(bool(acad))}-{int(bool(cert))}"
            for r in range(self._recognized_model.rowCount()):
                f2 = (self._recognized_model.index(r, 1).data(Qt.DisplayRole) or "").strip().lower()
                l2 = (self._recognized_model.index(r, 7).data(Qt.DisplayRole) or "").strip().lower()  # Badge Link
                d2 = (self._recognized_model.index(r, 4).data(Qt.DisplayRole) or "").strip()          # Recognition Date
                b2s= (self._recognized_model.index(r, 6).data(Qt.DisplayRole) or "").lower()          # Badge Types
                b2 = f"{int('academia' in b2s)}-{int('certified' in b2s)}"
                if f2 == full_l and (l2 == link_l or (d2 == dstr and b2 == bkey)):
                    return True
            return False
    
        board = new.get("board","")
        full  = new.get("full_name","")
        mail  = new.get("email","")
        addr  = new.get("address","")
        rdate = new.get("recognition_date","")
        link  = new.get("badge_link","")
    
        full_l = (full or "").strip().lower()
        link_l = (link or "").strip().lower()
        dstr   = (rdate or "").strip()
    
        rows_to_add = []
        for acad, cert in _badge_rows(new):
            if _dup_exists(full_l, link_l, dstr, acad, cert):
                continue
            try:
                dt = datetime.strptime(rdate, "%Y-%m-%d").date()
                vuntil = (dt + timedelta(days=365)).isoformat()
            except Exception:
                vuntil = ""
            badges = "; ".join([s for s,b in (("Academia", acad), ("Certified", cert)) if b])
            vals = [board, full, mail, addr, rdate, vuntil, badges, link]
            rows_to_add.append([QStandardItem(str(v)) for v in vals])
    
        if not rows_to_add:
            QMessageBox.information(self, "Duplicate", "This person/badge already exists.")
            return
    
        for items in rows_to_add:
            for it in items:
                it.setEditable(False)
            self._recognized_model.appendRow(items)
    
        self._recognized_fit_columns()
        self._recognized_apply_row_colors()
    
    def _recognized_edit(self) -> None:
        """
        Edit vybraného řádku.
        V EDIT režimu jsou badge typy i volba Sorted zdroje irelevantní (skryty).
        Výběr z tabulky mapujeme z proxy na zdrojový model.
        """
        from PySide6.QtWidgets import QMessageBox
        from PySide6.QtCore import Qt
        from datetime import datetime, timedelta
    
        sel = self.tbl_recognized.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "Edit", "Select a row to edit.")
            return
        # map proxy -> source row
        proxy_idx = sel[0]
        try:
            src_idx = self._rec_proxy.mapToSource(proxy_idx)
            r = src_idx.row()
        except Exception:
            r = proxy_idx.row()
    
        # Původní hodnoty
        cur = {
            "board": self._recognized_model.index(r,0).data(Qt.DisplayRole) or "",
            "full_name": self._recognized_model.index(r,1).data(Qt.DisplayRole) or "",
            "email": self._recognized_model.index(r,2).data(Qt.DisplayRole) or "",
            "address": self._recognized_model.index(r,3).data(Qt.DisplayRole) or "",
            "recognition_date": self._recognized_model.index(r,4).data(Qt.DisplayRole) or "",
            "badge_link": self._recognized_model.index(r,7).data(Qt.DisplayRole) or "",
        }
        btxt = (self._recognized_model.index(r,6).data(Qt.DisplayRole) or "").lower()
        cur["academia"]  = "academia"  in btxt
        cur["certified"] = "certified" in btxt
    
        upd = self._recognized_open_add_dialog(cur)
        if not upd:
            return
    
        # Přepočítej Valid Until = +365 dní
        try:
            dt = datetime.strptime(upd.get("recognition_date",""), "%Y-%m-%d").date()
            vuntil = (dt + timedelta(days=365)).isoformat()
        except Exception:
            vuntil = ""
    
        # Badge typy se v EDIT režimu nemění (ponecháme původní)
        acad_keep = cur["academia"]; cert_keep = cur["certified"]
        badges_txt = "; ".join([s for s,b in (("Academia", acad_keep), ("Certified", cert_keep)) if b])
    
        vals = [
            upd.get("board",""), upd.get("full_name",""), upd.get("email",""), upd.get("address",""),
            upd.get("recognition_date",""), vuntil, badges_txt, upd.get("badge_link",""),
        ]
        for c, v in enumerate(vals):
            self._recognized_model.setData(self._recognized_model.index(r, c), str(v))
    
        self._recognized_fit_columns()
        self._recognized_apply_row_colors()
        
    def _recognized_apply_row_colors(self) -> None:
        """
        Podbarvení řádků podle platnosti (Valid Until vs dnešek):
          - > 30 dnů: zelená
          - 0..30 dnů: žlutá
          - < 0 dnů: červená
        (Valid Until je ve sloupci index 5.)
        """
        from PySide6.QtGui import QColor, QBrush
        from PySide6.QtCore import Qt
        from datetime import datetime, date
    
        try:
            today = date.today()
            rows = self._recognized_model.rowCount()
            for r in range(rows):
                vuntil_s = self._recognized_model.index(r, 5).data(Qt.DisplayRole) or ""
                col_brush = None
                try:
                    vuntil = datetime.strptime(vuntil_s, "%Y-%m-%d").date()
                    days_left = (vuntil - today).days
                    if days_left > 30:
                        col_brush = QBrush(QColor(40, 140, 60, 60))    # zelená (jemná)
                    elif days_left >= 0:
                        col_brush = QBrush(QColor(200, 160, 20, 70))   # žlutá
                    else:
                        col_brush = QBrush(QColor(200, 60, 60, 80))    # červená
                except Exception:
                    col_brush = None
    
                if col_brush:
                    for c in range(self._recognized_model.columnCount()):
                        self._recognized_model.item(r, c).setBackground(col_brush)
                else:
                    for c in range(self._recognized_model.columnCount()):
                        self._recognized_model.item(r, c).setBackground(QBrush())
        except Exception:
            pass
    
    def _recognized_delete(self) -> None:
        """Smazání vybraných řádků (s potvrzením). Mapování výběru z proxy na zdrojový model."""
        from PySide6.QtWidgets import QMessageBox
        sel = self.tbl_recognized.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "Delete", "Select one or more rows to delete.")
            return
        if QMessageBox.question(self, "Delete", f"Delete {len(sel)} selected item(s)?") != QMessageBox.Yes:
            return
        # proxy -> source rows
        rows = []
        for pidx in sel:
            try:
                sidx = self._rec_proxy.mapToSource(pidx)
                rows.append(sidx.row())
            except Exception:
                rows.append(pidx.row())
        for r in sorted(set(rows), reverse=True):
            self._recognized_model.removeRow(r)

    def _recognized_fit_columns(self) -> None:
        """Do-fit sloupců po naplnění/změnách."""
        try:
            hdr = self.tbl_recognized.horizontalHeader()
            for c in range(self._recognized_model.columnCount()):
                self.tbl_recognized.resizeColumnToContents(c)
            hdr.setStretchLastSection(True)
        except Exception:
            pass

    # ----- Menu / actions -----
    def _build_menu(self) -> None:
        rescan_action = QAction("Rescan", self)
        rescan_action.triggered.connect(self.rescan)

        open_action = QAction("Open Selected PDF", self)
        open_action.triggered.connect(self.open_selected_pdf)

        export_csv_action = QAction("Export CSV (visible rows)…", self)
        export_csv_action.triggered.connect(self.export_csv)

        export_xlsx_action = QAction("Export XLSX (visible rows)…", self)
        export_xlsx_action.triggered.connect(self.export_xlsx)

        about_action = QAction("About", self)
        about_action.triggered.connect(self._about)

        # Minimal-change: actions directly on menubar (no new menus introduced)
        self.menuBar().addAction(rescan_action)
        self.menuBar().addAction(open_action)
        self.menuBar().addAction(export_csv_action)
        self.menuBar().addAction(export_xlsx_action)
        self.menuBar().addAction(about_action)
        
    def _gather_visible_records(self) -> list[PdfRecord]:
        model = self.table.model()
        if model is None:
            return []
        paths: list[str] = []
        FILE_COL = 16  # poslední sloupec po odebrání číslování
        if isinstance(model, QSortFilterProxyModel):
            src = model.sourceModel()
            if src is None:
                return []
            for r in range(model.rowCount()):
                pidx = model.index(r, FILE_COL)
                sidx = model.mapToSource(pidx)
                val = src.index(sidx.row(), FILE_COL).data(Qt.UserRole + 1)
                if val:
                    paths.append(str(val))
        else:
            for r in range(model.rowCount()):
                val = model.index(r, FILE_COL).data(Qt.UserRole + 1)
                if val:
                    paths.append(str(val))

        out: list[PdfRecord] = []
        for p in paths:
            rec = next((x for x in self.records if str(x.path) == p), None)
            if rec:
                out.append(rec)
        return out

    def _selected_record(self) -> Optional[PdfRecord]:
        sel = self.table.selectionModel()
        if not sel or not sel.hasSelection():
            return None
        index = sel.selectedRows()[0]
        FILE_COL = 16
        proxy = self.table.model()
        if isinstance(proxy, QSortFilterProxyModel):
            sidx = proxy.mapToSource(proxy.index(index.row(), FILE_COL))
            src = proxy.sourceModel()
        else:
            sidx = index
            src = self.table.model()
        path_str = src.index(sidx.row(), FILE_COL).data(Qt.UserRole + 1)
        for r in self.records:
            if str(r.path) == path_str:
                return r
        return None

    def _visible_columns(self) -> list[int]:
        """Return list of column indices currently visible in the table (proxy)."""
        model = self.table.model()
        if model is None:
            return []
        cols = model.columnCount()
        visible = []
        for c in range(cols):
            if not self.table.isColumnHidden(c):
                visible.append(c)
        return visible

    def export_csv(self) -> None:
        from PySide6.QtWidgets import QFileDialog
        import csv

        model = self.table.model()
        if model is None:
            QMessageBox.information(self, "Export CSV", "No data to export.")
            return

        visible_cols = self._visible_columns()
        if not visible_cols:
            QMessageBox.information(self, "Export CSV", "No visible columns to export.")
            return

        default_name = str((self.pdf_root / "export.csv"))
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV", default_name, "CSV Files (*.csv)")
        if not path:
            return

        # Build headers from visible columns (display text, newlines replaced)
        headers_list: list[str] = []
        for c in visible_cols:
            h = model.headerData(c, Qt.Horizontal, Qt.DisplayRole) or ""
            headers_list.append(str(h).replace("\n", " • "))

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(headers_list)
                for r in range(model.rowCount()):
                    row_vals = []
                    for c in visible_cols:
                        val = model.index(r, c).data(Qt.DisplayRole)
                        row_vals.append("" if val is None else str(val))
                    writer.writerow(row_vals)
            QMessageBox.information(self, "Export CSV", f"Exported {model.rowCount()} rows to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export CSV", f"Failed to write CSV:\n{e}")

    def export_xlsx(self) -> None:
        from PySide6.QtWidgets import QFileDialog
        try:
            from openpyxl import Workbook
        except Exception:
            QMessageBox.critical(
                self, "Export XLSX",
                "The 'openpyxl' package is required for XLSX export.\n"
                "Install it in your environment:\n\npip install openpyxl"
            )
            return

        model = self.table.model()
        if model is None:
            QMessageBox.information(self, "Export XLSX", "No data to export.")
            return

        visible_cols = self._visible_columns()
        if not visible_cols:
            QMessageBox.information(self, "Export XLSX", "No visible columns to export.")
            return

        default_name = str((self.pdf_root / "export.xlsx"))
        path, _ = QFileDialog.getSaveFileName(self, "Save XLSX", default_name, "Excel Workbook (*.xlsx)")
        if not path:
            return

        headers_list: list[str] = []
        for c in visible_cols:
            h = model.headerData(c, Qt.Horizontal, Qt.DisplayRole) or ""
            headers_list.append(str(h).replace("\n", " • "))

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ISTQB Applications"
            ws.append(headers_list)
            for r in range(model.rowCount()):
                row_vals = []
                for c in visible_cols:
                    val = model.index(r, c).data(Qt.DisplayRole)
                    row_vals.append("" if val is None else str(val))
                ws.append(row_vals)
            # Auto width
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)
            wb.save(path)
            QMessageBox.information(self, "Export XLSX", f"Exported {model.rowCount()} rows to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export XLSX", f"Failed to write XLSX:\n{e}")
            
    def _about(self) -> None:
        QMessageBox.information(
            self, "About",
            "ISTQB Academia PDF Aggregator\n\n"
            "Scans the 'PDF' folder next to the script and aggregates key fields.\n"
            "UI: English • Theme: Dark • Target: macOS"
        )

    def export_selected_to_sorted(self) -> None:
        from PySide6.QtWidgets import QMessageBox
        from pathlib import Path
        import shutil
    
        sel = self.table.selectionModel()
        if not sel or not sel.hasSelection():
            QMessageBox.information(self, "Export", "Please select at least one row.")
            return
    
        # Detekce sloupců Board a File name (předpoklad: poslední sloupec = File name s fullpath v UserRole+1)
        model = self.table.model()
        if model is None:
            return
    
        # Zjisti indexy vybraných řádků v source modelu
        rows = [i.row() for i in sel.selectedRows()]
        if not rows:
            QMessageBox.information(self, "Export", "Není co exportovat.")
            return
    
        # Najdi sloupce
        COL_BOARD = 0
        COL_FILE  = model.columnCount() - 1
    
        exported = 0
        for r in rows:
            board = (model.index(r, COL_BOARD).data() or "").strip()
            # Full path je v UserRole + 1
            full = model.index(r, COL_FILE).data(Qt.UserRole + 1) or model.index(r, COL_FILE).data()
            if not full:
                continue
            full = Path(str(full))
            if not full.exists():
                continue
    
            # Cílová složka Sorted PDFs/<board> (board může být prázdný → kořen Sorted PDFs)
            dest_dir = (self.sorted_root / board).resolve()
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / full.name
    
            try:
                shutil.copy2(full, dest)
                exported += 1
            except Exception:
                continue
    
        # Rescan Sorted (naparsuje a upsertne do DB)
        self.rescan_sorted()
    
        try:
            self.statusBar().showMessage(f"Exportováno do 'Sorted PDFs': {exported} souborů.")
        except Exception:
            pass

    # ----- Overview tab -----
    def _build_overview_tab(self) -> None:
        from PySide6.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QLineEdit, QPushButton,
            QTableView, QToolButton, QMenu, QCheckBox, QStyle, QApplication,
            QStyledItemDelegate, QStyleOptionViewItem
        )
        from PySide6.QtGui import QStandardItemModel, QIcon, QPainter
        from PySide6.QtCore import Qt, QTimer
    
        layout = QVBoxLayout()
        controls = QHBoxLayout()
    
        # Unparsed
        self.btn_unparsed = QToolButton(self)
        self.btn_unparsed.setText("Unparsed")
        self.btn_unparsed.setToolTip("Show PDFs found on disk that are not present in Overview")
        self.btn_unparsed.setIcon(self.style().standardIcon(QStyle.SP_MessageBoxWarning))
        self.btn_unparsed.setAutoRaise(True)
        self.btn_unparsed.setStyleSheet("QToolButton { color: #ff6b6b; font-weight: 600; }")
        self.btn_unparsed.clicked.connect(self.show_unparsed_report)
    
        # Export (Overview)
        self.btn_export = QToolButton(self)
        self.btn_export.setToolTip("Export…")
        self.btn_export.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        self.btn_export.setAutoRaise(True)
        self.btn_export.clicked.connect(self.on_export_overview)
        controls.addWidget(self.btn_export)
    
        # Board combobox – model naplníme ve _rebuild_board_combo()
        self.board_combo = QComboBox()
        self.board_combo.currentTextChanged.connect(self._filter_board)
    
        # Fulltext
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search…")
        self.search_edit.textChanged.connect(self._filter_text)
    
        # Open Selected PDF
        self.open_btn = QPushButton("Open Selected PDF")
        self.open_btn.clicked.connect(self.open_selected_pdf)
    
        # Checkbox Sorted (filtr ve view)
        self.chk_overview_sorted = QCheckBox("Sorted")
        self.chk_overview_sorted.setToolTip("Show/hide records that are already in 'Sorted PDFs'")
        self.chk_overview_sorted.setChecked(True)
        self.chk_overview_sorted.toggled.connect(self._on_overview_sorted_toggled)
    
        # Controls layout
        controls.addWidget(self.btn_unparsed)
        controls.addSpacing(12)
        controls.addWidget(QLabel("Board:"))
        controls.addWidget(self.board_combo, 1)
        controls.addSpacing(12)
        controls.addWidget(QLabel("Search:"))
        controls.addWidget(self.search_edit, 4)
        controls.addSpacing(12)
        controls.addWidget(self.chk_overview_sorted)
        controls.addSpacing(12)
        controls.addWidget(self.open_btn)
        layout.addLayout(controls)
    
        # Tabulka Overview
        self.table = QTableView()
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setSelectionMode(QTableView.ExtendedSelection)
        self.table.doubleClicked.connect(self.open_selected_pdf)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setMinimumHeight(44)
    
        # Hlavičky – poslední sloupec Sorted
        self._headers = [
            "Board",
            "Application\nApplication Type",
            "Name of Your Academic Institution\nInstitution Name",
            "Name of Your Academic Institution\nCandidate Name",
            "Wished Recognitions\nAcademia Recognition",
            "Wished Recognitions\nCertified Recognition",
            "Contact details for Information exchange\nFull Name",
            "Contact details for Information exchange\nEmail Address",
            "Contact details for Information exchange\nPhone Number",
            "Contact details for Information exchange\nPostal Address",
            "Eligibility Evidence\nSyllabi Integration",
            "Eligibility Evidence\nCourses/Modules",
            "Eligibility Evidence\nProof of ISTQB Certifications",
            "Eligibility Evidence\nUniversity Links",
            "Eligibility Evidence\nAdditional Info/Documents",
            "Signature Date",
            "File\nFile name",
            "Sorted",
        ]
    
        # Model + proxy
        if not hasattr(self, "_source_model"):
            self._source_model = QStandardItemModel(0, len(self._headers), self)
            self._source_model.setHorizontalHeaderLabels(self._headers)
        else:
            self._source_model.setColumnCount(len(self._headers))
            self._source_model.setHorizontalHeaderLabels(self._headers)
    
        if not hasattr(self, "_proxy"):
            self._proxy = RecordsModel(self._headers, self)
            self._proxy.setSourceModel(self._source_model)
            self._proxy.setDynamicSortFilter(True)
        self.table.setModel(self._proxy)
    
        # Delegát pro centrování ikon
        class IconCenterDelegate(QStyledItemDelegate):
            def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
                icon = index.data(Qt.DecorationRole)
                if icon:
                    opt = QStyleOptionViewItem(option)
                    self.initStyleOption(opt, index)
                    txt, ico = opt.text, opt.icon
                    opt.text, opt.icon = "", QIcon()
                    style = option.widget.style() if option.widget else QApplication.style()
                    style.drawControl(QStyle.CE_ItemViewItem, opt, painter, option.widget)
                    pm = icon.pixmap(opt.decorationSize if opt.decorationSize.isValid() else opt.rect.size())
                    x = opt.rect.x() + (opt.rect.width() - pm.width()) // 2
                    y = opt.rect.y() + (opt.rect.height() - pm.height()) // 2
                    painter.drawPixmap(x, y, pm)
                else:
                    super().paint(painter, option, index)
    
        # Hiding delegát pro Board
        self.table.setItemDelegateForColumn(0, BoardHidingDelegate(self.table))
    
        # Najdi indexy sloupců
        def _find_col_tail(tail: str) -> int | None:
            for i, h in enumerate(self._headers):
                t = h.split("\n")[-1].strip() if "\n" in h else h.strip()
                if t.lower() == tail.lower():
                    return i
            return None
    
        col_acad = _find_col_tail("Academia Recognition")
        col_cert = _find_col_tail("Certified Recognition")
        col_sorted = _find_col_tail("Sorted")
        center_delegate = IconCenterDelegate(self.table)
        if col_acad is not None:
            self.table.setItemDelegateForColumn(col_acad, center_delegate)
        if col_cert is not None:
            self.table.setItemDelegateForColumn(col_cert, center_delegate)
        if col_sorted is not None:
            self.table.setItemDelegateForColumn(col_sorted, center_delegate)
    
        # Skryj Eligibility sloupce
        for c in (10, 11, 12, 13, 14):
            self.table.setColumnHidden(c, True)
    
        # Kontextové menu – export do Sorted (po exportu jednorázově zreviduj Sorted sloupec)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        def _ctx(pos):
            idx = self.table.indexAt(pos)
            if idx.isValid():
                if not self.table.selectionModel().isSelected(idx):
                    self.table.selectRow(idx.row())
            menu = QMenu(self.table)
            act_export_sorted = menu.addAction("Exportovat do složky 'Sorted PDFs'")
            chosen = menu.exec_(self.table.viewport().mapToGlobal(pos))
            if chosen == act_export_sorted:
                self.export_selected_to_sorted()
                # minimální doplněk: po exportu přepočti Sorted + aplikuj hiding (tiché, bez zásahu do výběru)
                try:
                    self._overview_update_sorted_flags()
                    self._overview_apply_sorted_row_hiding()
                except Exception:
                    pass
        if not hasattr(self, "_overview_ctx_connected"):
            self.table.customContextMenuRequested.connect(_ctx)
            self._overview_ctx_connected = True
    
        layout.addWidget(self.table, 1)
        self.overview_tab.setLayout(layout)
    
        # Po vystavění: jednorázově spočti Sorted a rebuildni Board combobox
        def _post_build():
            try:
                self._overview_update_sorted_flags()   # jen jednou
                self._overview_apply_sorted_row_hiding()
                self._rebuild_board_combo()
            except Exception:
                pass
        QTimer.singleShot(0, _post_build)
    
        # Při změně řazení pouze aplikuj hiding (nepřepočítávej Sorted)
        try:
            self.table.horizontalHeader().sortIndicatorChanged.connect(lambda *_: self._overview_apply_sorted_row_hiding())
        except Exception:
            pass
        
    def _overview_find_col(self, tail: str) -> int | None:
        """
        Najde index sloupce podle 'tail' = poslední řádek popisku (po \n).
        Např. tail="File name" → vrátí index sloupce s "File\nFile name".
        """
        headers = getattr(self, "_headers", [])
        for i, h in enumerate(headers):
            t = h.split("\n")[-1].strip() if "\n" in h else h.strip()
            if t.lower() == tail.lower():
                return i
        return None
    
    def _collect_sorted_filenames(self) -> set[str]:
        """
        Vrátí množinu názvů souborů, které jsou považovány za 'Sorted'.
        1) in-memory DB (self.sorted_db / podobné),
        2) fallback: existující soubory v adresáři 'Sorted PDFs' (pokud jej najdeme).
        """
        from pathlib import Path
        names: set[str] = set()
    
        # 1) In-memory
        for attr in ("sorted_db", "_sorted_db", "sorted_records", "_sorted_records", "sorted_data", "_sorted_data"):
            cnt = getattr(self, attr, None)
            if not cnt:
                continue
            def _iter(c):
                if isinstance(c, dict):
                    for v in c.values():
                        yield v
                elif isinstance(c, list):
                    for v in c:
                        yield v
            for rec in _iter(cnt):
                try:
                    # preferuj path.name
                    fn = None
                    p = getattr(rec, "path", None)
                    if p:
                        try:
                            fn = Path(p).name
                        except Exception:
                            fn = None
                    if not fn and isinstance(rec, dict):
                        fn = rec.get("file_name") or (Path(rec.get("path")).name if rec.get("path") else None)
                    if fn:
                        names.add(str(fn))
                except Exception:
                    pass
            if names:
                break
    
        # 2) Fallback: projdi 'Sorted PDFs'
        try:
            roots = []
            # zkus odvodit root
            if hasattr(self, "pdf_root") and self.pdf_root:
                roots.append(Path(self.pdf_root))
            roots.append(Path(__file__).resolve().parents[2])  # repo root
            for root in roots:
                d = root / "Sorted PDFs"
                if d.exists() and d.is_dir():
                    for p in d.glob("**/*.pdf"):
                        names.add(p.name)
                    break
        except Exception:
            pass
    
        return names
    
    def _overview_update_sorted_flags(self) -> None:
        """
        Dosadí do POSLEDNÍHO sloupce 'Sorted' hodnotu 'Yes'/'' + ikonku (DecorationRole),
        centrování (TextAlignmentRole) a světle šedé pozadí (BackgroundRole).
        Tichý in-place update: NEMĚNÍ výběr řádků.
        """
        from PySide6.QtCore import Qt
        from PySide6.QtWidgets import QStyle
        from PySide6.QtGui import QBrush, QColor
    
        # sloupce
        fn_col = self._overview_find_col("File name")
        sorted_col = self._overview_find_col("Sorted")
        if fn_col is None or sorted_col is None:
            return
        if not hasattr(self, "_source_model"):
            return
    
        # připrav hashe Sorted a zdroje pro roli
        sorted_hashes = self._collect_sorted_hashes()
        icon_ok = self.style().standardIcon(QStyle.SP_DialogApplyButton)
        bg_brush = QBrush(QColor(240, 240, 240))  # světlá šedá
    
        rows = self._source_model.rowCount()
        for r in range(rows):
            idx_fn = self._source_model.index(r, fn_col)
            fname = (self._source_model.data(idx_fn, Qt.DisplayRole) or "").strip()
    
            mark = ""
            try:
                path = self._find_record_path_for_filename(fname)
                if path:
                    dig = self._hash_file(path)
                    if dig and dig in sorted_hashes:
                        mark = "Yes"
            except Exception:
                mark = ""
    
            idx_sorted = self._source_model.index(r, sorted_col)
            # nastav pouze data/role (žádné zásahy do selection)
            if (self._source_model.data(idx_sorted, Qt.DisplayRole) or "") != mark:
                self._source_model.setData(idx_sorted, mark, Qt.DisplayRole)
            self._source_model.setData(idx_sorted, Qt.AlignCenter, Qt.TextAlignmentRole)
            self._source_model.setData(idx_sorted, icon_ok if mark == "Yes" else None, Qt.DecorationRole)
            self._source_model.setData(idx_sorted, bg_brush, Qt.BackgroundRole)
            
    def _overview_apply_sorted_row_hiding(self) -> None:
        """
        Skryje/ukáže řádky podle checkboxu 'Sorted'.
        Nepřepočítává Sorted; pouze čte DisplayRole v proxy.
        """
        from PySide6.QtCore import Qt
    
        if not hasattr(self, "table") or not hasattr(self, "_proxy"):
            return
        sorted_col = self._overview_find_col("Sorted")
        if sorted_col is None:
            return
    
        show_sorted = bool(self.chk_overview_sorted.isChecked())
        rows = self._proxy.rowCount()
        for r in range(rows):
            idx = self._proxy.index(r, sorted_col)
            val = self._proxy.data(idx, Qt.DisplayRole) or ""
            self.table.setRowHidden(r, (val == "Yes") and (not show_sorted))
        
    def _on_overview_sorted_toggled(self, checked: bool) -> None:
        """
        Reakce na checkbox 'Sorted' – pouze přepočítá hiding dle aktuálních dat.
        """
        try:
            self._overview_apply_sorted_row_hiding()
        except Exception:
            pass
        
    def _find_record_path_for_filename(self, fname: str):
        """
        Zjisti Path k PDF pro Overview řádek:
          1) pokud v self.records existuje rec.path s tímto názvem,
          2) pokud 'fname' v buňce je už plná cesta a existuje,
          3) rekurzivně vyhledej pod pdf_root (case-insensitive podle názvu).
        Cache: self._fname_to_path_cache
        """
        from pathlib import Path
    
        if not fname:
            return None
        if not hasattr(self, "_fname_to_path_cache"):
            self._fname_to_path_cache = {}
    
        if fname in self._fname_to_path_cache:
            return self._fname_to_path_cache[fname]
    
        # plná cesta v buňce?
        try:
            p = Path(fname)
            if p.exists() and p.is_file():
                self._fname_to_path_cache[fname] = p
                return p
        except Exception:
            pass
    
        # záznamy -> path
        try:
            for rec in (getattr(self, "records", None) or []):
                p = getattr(rec, "path", None)
                if p:
                    pp = Path(p)
                    if pp.name == fname or str(pp) == fname:
                        self._fname_to_path_cache[fname] = pp
                        return pp
        except Exception:
            pass
    
        # vyhledat pod pdf_root
        try:
            root = getattr(self, "pdf_root", None)
            if root:
                root = Path(root)
                target = fname.lower()
                for pp in root.rglob("*.pdf"):
                    try:
                        if pp.name.lower() == target:
                            self._fname_to_path_cache[fname] = pp
                            return pp
                    except Exception:
                        continue
        except Exception:
            pass
    
        self._fname_to_path_cache[fname] = None
        return None
    
    def _enumerate_all_pdfs(self) -> list[str]:
        """
        Vrátí list absolutních cest na *.pdf pod self.pdf_root (rekurzivně),
        s ignorováním jakékoli složky '__archive__'.
        """
        from pathlib import Path
        root = getattr(self, "pdf_root", None)
        if not root:
            return []
        root = Path(root).resolve()
        out: list[str] = []
        try:
            for p in root.rglob("*.pdf"):
                try:
                    rel = p.resolve().relative_to(root)
                    if "__archive__" in rel.parts:
                        continue
                    out.append(str(p.resolve()))
                except Exception:
                    out.append(str(p.resolve()))
        except Exception:
            pass
        return out
        
    def _collect_sorted_hashes(self) -> set[str]:
        """
        Vrátí set SHA-256 hashů všech PDF ve složce(ách) 'Sorted PDFs' (rekurzivně).
        """
        hashes: set[str] = set()
        try:
            dirs = self._find_sorted_dirs()
            for d in dirs:
                try:
                    for p in d.rglob("*.pdf"):
                        dig = self._hash_file(p)
                        if dig:
                            hashes.add(dig)
                    break  # stačí první nalezená složka
                except Exception:
                    continue
        except Exception:
            pass
        return hashes
    
    def _find_sorted_dirs(self):
        """
        Najdi složky 'Sorted PDFs' (case-insensitive) pod pdf_root i pod repo rootem.
        Vrací list Path; první nalezená se použije v _collect_sorted_hashes().
        """
        from pathlib import Path
        cand = []
        names = {"sorted pdfs", "sorted_pdfs", "sorted-pdfs"}
        roots = []
        try:
            if getattr(self, "pdf_root", None):
                roots.append(Path(self.pdf_root))
            roots.append(Path(__file__).resolve().parents[2])  # repo root
        except Exception:
            pass
    
        for root in roots:
            try:
                direct = root / "Sorted PDFs"
                if direct.exists() and direct.is_dir():
                    cand.append(direct)
                    continue
                # jedno patro dolů (case-insensitive)
                for d in [root] + [p for p in root.iterdir() if p.is_dir()]:
                    for sub in d.iterdir():
                        if sub.is_dir() and sub.name.lower() in names:
                            cand.append(sub)
                            break
                    if cand:
                        break
            except Exception:
                continue
            if cand:
                break
        return cand
        
    def _hash_file(self, path) -> str:
        """
        SHA-256 daného souboru (čtení po 1 MiB blocích). Výsledek cache-uje.
        """
        from pathlib import Path
        import hashlib
    
        try:
            p = Path(path)
            if not hasattr(self, "_hash_cache"):
                self._hash_cache = {}
            key = str(p.resolve())
            if key in self._hash_cache:
                return self._hash_cache[key]
            h = hashlib.sha256()
            with p.open("rb") as fh:
                for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                    h.update(chunk)
            dig = h.hexdigest()
            self._hash_cache[key] = dig
            return dig
        except Exception:
            return ""
        
    def _rebuild_board_combo(self) -> None:
        """
        Sestaví combobox se dvěma sekcemi:
          - All
          - [boards přítomné v Overview]
          - (separator – disabled item)
          - [zbylé KNOWN_BOARDS]
        Zachová aktuální výběr, pokud je to možné.
        """
        from PySide6.QtGui import QStandardItemModel, QStandardItem
        from PySide6.QtCore import Qt
    
        try:
            prev = self.board_combo.currentText()
        except Exception:
            prev = "All"
    
        present: set[str] = set()
        try:
            rows = self._source_model.rowCount()
            for r in range(rows):
                val = self._source_model.index(r, 0).data(Qt.DisplayRole) or ""
                if val:
                    present.add(val)
        except Exception:
            pass
    
        known_sorted = sorted(KNOWN_BOARDS)
        present_sorted = sorted(present)
        rest = [b for b in known_sorted if b not in present]
    
        m = QStandardItemModel(self.board_combo)
        def _add_txt(txt: str, enabled: bool = True):
            it = QStandardItem(txt)
            if not enabled:
                it.setFlags(Qt.NoItemFlags)
            m.appendRow(it)
    
        _add_txt("All")
        for b in present_sorted:
            _add_txt(b)
        # separator jen pokud jsou nějaké přítomné a zároveň existuje zbytek
        if present_sorted and rest:
            _add_txt("────────────", enabled=False)
        for b in rest:
            _add_txt(b)
    
        self.board_combo.setModel(m)
    
        # obnov výběr
        # pokud předchozí neexistuje (separator), přepni na All
        idx = None
        for row in range(m.rowCount()):
            if m.item(row).flags() & Qt.ItemIsEnabled:
                if m.item(row).text() == prev:
                    idx = row
                    break
        if idx is None:
            idx = 0
        self.board_combo.setCurrentIndex(idx)

    def _enumerate_all_pdfs(self) -> list[str]:
        """
        Vrátí list absolutních cest na *.pdf pod self.pdf_root (rekurzivně),
        s ignorováním jakékoli složky '__archive__'.
        """
        from pathlib import Path
        root = getattr(self, "pdf_root", None)
        if not root:
            return []
        root = Path(root).resolve()
        out: list[str] = []
        try:
            for p in root.rglob("*.pdf"):
                try:
                    rel = p.resolve().relative_to(root)
                    if "__archive__" in rel.parts:
                        continue
                    out.append(str(p.resolve()))
                except Exception:
                    # pokud by nešla relativní cesta, stále přidáme (bez filtru)
                    out.append(str(p.resolve()))
        except Exception:
            pass
        return out

    def show_unparsed_report(self) -> None:
        """
        Ad-hoc audit: porovná PDF v self.pdf_root (rekurzivně; ignoruje '__archive__')
        s aktuálně zobrazenými záznamy (self.records).
        Zobrazí dialog s přehledem "unparsed" PDF (Board, File name, Full path).
        """
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QLabel,
            QTreeWidget, QTreeWidgetItem, QPushButton
        )
        from PySide6.QtCore import Qt
        from pathlib import Path
    
        # 1) PDF na disku (mimo __archive__)
        all_pdfs = self._enumerate_all_pdfs()
    
        # 2) PDF v Overview (už naparsovaná) – vezmeme absolutní cesty
        parsed_paths = set()
        try:
            for r in getattr(self, "records", []):
                p = getattr(r, "path", None)
                if p:
                    parsed_paths.add(str(Path(p).resolve()))
        except Exception:
            pass
    
        # 3) Rozdíl = unparsed
        unparsed = []
        for p in all_pdfs:
            ap = str(Path(p).resolve())
            if ap not in parsed_paths:
                # board = první složka pod pdf_root (pokud existuje)
                try:
                    rel = Path(p).resolve().relative_to(self.pdf_root.resolve())
                    board = rel.parts[0] if len(rel.parts) > 1 else ""
                except Exception:
                    board = ""
                unparsed.append((board, Path(p).name, ap))
    
        # 4) Dialog s výsledky
        dlg = QDialog(self)
        dlg.setWindowTitle("Unparsed PDFs")
        dlg.resize(800, 500)
    
        lay = QVBoxLayout(dlg)
        header = QLabel(f"PDFs on disk not present in Overview: {len(unparsed)}")
        header.setStyleSheet("QLabel { color: #ff6b6b; font-weight: 600; }")
        lay.addWidget(header)
    
        tree = QTreeWidget()
        tree.setHeaderLabels(["Board", "File name", "Full path"])
        tree.header().setDefaultAlignment(Qt.AlignCenter)
        tree.header().setStretchLastSection(True)
        for board, fname, fullp in sorted(unparsed, key=lambda t: (t[0].lower(), t[1].lower())):
            it = QTreeWidgetItem([board or "—", fname, fullp])
            tree.addTopLevelItem(it)
        tree.expandAll()
        lay.addWidget(tree, 1)
    
        # Tlačítka
        btns = QHBoxLayout()
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dlg.accept)
        btns.addStretch(1)
        btns.addWidget(btn_close)
        lay.addLayout(btns)
    
        # 5) Aktualizuj badge na tlačítku
        try:
            if len(unparsed) > 0:
                self.btn_unparsed.setText(f"Unparsed: {len(unparsed)}")
                self.statusBar().showMessage(f"Unparsed PDFs: {len(unparsed)}")
            else:
                self.btn_unparsed.setText("Unparsed")
                self.statusBar().showMessage("All PDFs present in Overview.")
        except Exception:
            pass
    
        dlg.exec()

    def _collect_available_boards(self) -> list[str]:
        """
        Získá boardy, pro které skutečně existují PDF ve složce PDF/<board>/...
        Ignoruje podsložku '__archive__'. Pokud není pdf_root dostupný,
        vrátí unikátní boardy z již naparsovaných self.records.
        """
        from pathlib import Path
        boards: set[str] = set()
    
        root = getattr(self, "pdf_root", None)
        if isinstance(root, Path) and root.exists():
            try:
                # první úroveň podsložek pod PDF je board
                for sub in root.iterdir():
                    if not sub.is_dir():
                        continue
                    if sub.name == "__archive__":
                        continue
                    # zkontroluj aspoň jedno .pdf uvnitř (rekurzivně)
                    found_pdf = False
                    for p in sub.rglob("*.pdf"):
                        try:
                            rel = p.relative_to(root)
                            if "__archive__" in rel.parts:
                                continue
                            found_pdf = True
                            break
                        except Exception:
                            continue
                    if found_pdf:
                        boards.add(sub.name)
            except Exception:
                pass
    
        if not boards:
            try:
                for r in getattr(self, "records", []):
                    if getattr(r, "board", None):
                        boards.add(r.board)
            except Exception:
                pass
    
        return sorted(boards)
    
    def on_export_overview(self) -> None:
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QCheckBox, QLabel,
            QComboBox, QPushButton, QListWidget, QListWidgetItem, QScrollArea,
            QWidget, QGridLayout, QDialogButtonBox, QFileDialog, QMessageBox, QRadioButton
        )
        from PySide6.QtCore import Qt
        from datetime import datetime
        import os
    
        # FIELDS v pořadí Overview (včetně nových)
        FIELDS: list[tuple[str, str]] = [
            ("Board", "board"),
            ("Application Type", "application_type"),
            ("Institution Name", "institution_name"),
            ("Candidate Name", "candidate_name"),
            ("Academia Recognition", "recognition_academia"),
            ("Certified Recognition", "recognition_certified"),
            ("Full Name", "contact_full_name"),
            ("Email Address", "contact_email"),
            ("Phone Number", "contact_phone"),
            ("Postal Address", "contact_postal_address"),
            ("Syllabi Integration", "syllabi_integration_description"),
            ("Courses/Modules", "courses_modules_list"),
            ("Proof of ISTQB Certifications", "proof_of_istqb_certifications"),
            ("University Links", "university_links"),
            ("Additional Info/Documents", "additional_information_documents"),
            ("Printed Name, Title", "printed_name_title"),
            ("Signature Date", "signature_date"),
            ("Receiving Member Board", "receiving_member_board"),
            ("Date Received", "date_received"),
            ("Validity Start Date", "validity_start_date"),
            ("Validity End Date", "validity_end_date"),
            ("File name", "file_name"),
        ]
        # --- pomocné: zjisti počet vybraných řádků + sadu vybraných názvů souborů z tabulky ---
        selected_count = 0
        selected_file_names: set[str] = set()
        try:
            view = getattr(self, "table", None)
            model = view.model() if view else None
            selm = view.selectionModel() if view else None
    
            # najdi index sloupce "File name" z self._headers (poslední řádek popisku po \n)
            col_file = None
            try:
                for i, h in enumerate(getattr(self, "_headers", [])):
                    lbl = h.split("\n")[-1].strip() if "\n" in h else h.strip()
                    if lbl.lower() == "file name":
                        col_file = i
                        break
                if col_file is None:
                    # fallback: první sloupec obsahující "File"
                    for i, h in enumerate(getattr(self, "_headers", [])):
                        if "File" in h:
                            col_file = i
                            break
            except Exception:
                col_file = None
    
            if view and model and selm and col_file is not None:
                for idx in selm.selectedRows(col_file):
                    val = model.data(idx, Qt.DisplayRole)
                    if val:
                        selected_file_names.add(str(val))
            selected_count = len(selected_file_names)
        except Exception:
            selected_count = 0
            selected_file_names = set()
    
        # --- dialog ---
        class ExportDialog(QDialog):
            def __init__(self, boards_avail: list[str], parent=None):
                super().__init__(parent)
                self.setWindowTitle("Export options")
                main = QVBoxLayout(self)
    
                # --- Formáty ---
                grp_fmt = QGroupBox("Export formats")
                fmt_lay = QHBoxLayout(grp_fmt)
                self.chk_csv = QCheckBox("CSV")
                self.chk_xlsx = QCheckBox("XLSX")
                self.chk_txt = QCheckBox("TXT (formatted)")
                # default: CSV + TXT zapnuto, XLSX vypnuto (bez nové závislosti)
                self.chk_csv.setChecked(True)
                self.chk_txt.setChecked(True)
                self.chk_xlsx.setChecked(False)
                fmt_lay.addWidget(self.chk_csv)
                fmt_lay.addWidget(self.chk_xlsx)
                fmt_lay.addWidget(self.chk_txt)
                main.addWidget(grp_fmt)
    
                # --- Scope (NOVÉ) ---
                grp_scope = QGroupBox("Scope")
                scope_lay = QVBoxLayout(grp_scope)
                self.rb_selected = QRadioButton(f"Selected rows ({selected_count})")
                self.rb_all = QRadioButton("All rows")
                if selected_count > 0:
                    self.rb_selected.setChecked(True)
                else:
                    self.rb_selected.setEnabled(False)
                    self.rb_all.setChecked(True)
                scope_lay.addWidget(self.rb_selected)
                scope_lay.addWidget(self.rb_all)
                main.addWidget(grp_scope)
    
                # --- Boardy ---
                grp_b = QGroupBox("Boards to export")
                b_lay = QGridLayout(grp_b)
                self.all_boards = QCheckBox("All")
                self.all_boards.setChecked(True)
                b_lay.addWidget(self.all_boards, 0, 0, 1, 2)
    
                self.cmb_board = QComboBox()
                self.cmb_board.addItems(boards_avail)
                self.btn_add = QPushButton("Add")
                b_lay.addWidget(QLabel("Board:"), 1, 0)
                b_lay.addWidget(self.cmb_board, 1, 1)
                b_lay.addWidget(self.btn_add, 1, 2)
    
                self.list_sel = QListWidget()
                b_lay.addWidget(QLabel("Selected boards:"), 2, 0, 1, 3)
                b_lay.addWidget(self.list_sel, 3, 0, 1, 3)
                self.btn_remove = QPushButton("Remove selected")
                b_lay.addWidget(self.btn_remove, 4, 0, 1, 3)
    
                def _toggle_board_ui():
                    enabled = not self.all_boards.isChecked()
                    self.cmb_board.setEnabled(enabled)
                    self.btn_add.setEnabled(enabled)
                    self.list_sel.setEnabled(enabled)
                    self.btn_remove.setEnabled(enabled)
    
                self.all_boards.toggled.connect(_toggle_board_ui)
                self.btn_add.clicked.connect(lambda: self._add_board())
                self.btn_remove.clicked.connect(lambda: self._remove_sel())
                _toggle_board_ui()
                main.addWidget(grp_b)
    
                # --- Pole ---
                grp_fields = QGroupBox("Fields to export")
                f_lay = QVBoxLayout(grp_fields)
                self.chk_all_fields = QCheckBox("Select all")
                self.chk_all_fields.setChecked(True)
                f_lay.addWidget(self.chk_all_fields)
    
                self.scroll = QScrollArea()
                self.scroll.setWidgetResizable(True)
                cont = QWidget()
                cont_lay = QVBoxLayout(cont)
                self.field_checks: list[tuple[str, str, QCheckBox]] = []
                for label, key in FIELDS:
                    cb = QCheckBox(label)
                    cb.setChecked(True)
                    self.field_checks.append((label, key, cb))
                    cont_lay.addWidget(cb)
                cont_lay.addStretch(1)
                self.scroll.setWidget(cont)
                f_lay.addWidget(self.scroll)
    
                def _toggle_all():
                    state = self.chk_all_fields.isChecked()
                    for _, _, cb in self.field_checks:
                        cb.setChecked(state)
                self.chk_all_fields.toggled.connect(_toggle_all)
                main.addWidget(grp_fields)
    
                # --- tlačítka ---
                btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
                main.addWidget(btns)
                btns.accepted.connect(self._accept)
                btns.rejected.connect(self.reject)
    
                self.result = None  # dict result
    
            def _add_board(self):
                txt = self.cmb_board.currentText().strip()
                if not txt:
                    return
                # nepřidávej duplicitně
                for i in range(self.list_sel.count()):
                    if self.list_sel.item(i).text() == txt:
                        return
                self.list_sel.addItem(QListWidgetItem(txt))
    
            def _remove_sel(self):
                for it in self.list_sel.selectedItems():
                    self.list_sel.takeItem(self.list_sel.row(it))
    
            def _accept(self):
                # formáty
                fmts = []
                if self.chk_csv.isChecked(): fmts.append("csv")
                if self.chk_xlsx.isChecked(): fmts.append("xlsx")
                if self.chk_txt.isChecked(): fmts.append("txt")
                if not fmts:
                    QMessageBox.warning(self, "Export", "Select at least one format.")
                    return
    
                # scope
                scope = "selected" if self.rb_selected.isChecked() else "all"
    
                # boardy
                boards = None
                if not self.all_boards.isChecked():
                    boards = []
                    for i in range(self.list_sel.count()):
                        boards.append(self.list_sel.item(i).text())
                    if not boards:
                        QMessageBox.warning(self, "Export", "Add at least one board or check 'All'.")
                        return
    
                # pole
                fields = [(lab, key) for (lab, key, cb) in self.field_checks if cb.isChecked()]
                if not fields:
                    QMessageBox.warning(self, "Export", "Select at least one field.")
                    return
    
                self.result = {
                    "formats": fmts,
                    "boards": boards,   # None = All
                    "fields": fields,   # list of (label, key)
                    "scope": scope,     # "selected" | "all"
                }
                self.accept()
    
        # připrav data pro dialog
        boards_avail = self._collect_available_boards()
        dlg = ExportDialog(boards_avail, self)
        if dlg.exec_() != QDialog.Accepted or dlg.result is None:
            return
    
        formats: list[str] = dlg.result["formats"]
        boards_sel: list[str] | None = dlg.result["boards"]
        fields: list[tuple[str, str]] = dlg.result["fields"]
        scope: str = dlg.result.get("scope", "all")
    
        # výběr cesty (jeden dialog – základní soubor, ostatní formáty vedle)
        # navržený název podle času a případných boardů
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        suffix = "all" if boards_sel is None else ("+".join(boards_sel[:3]) + ("+more" if boards_sel and len(boards_sel) > 3 else ""))
        base_name = f"export-{suffix}-{ts}".replace(" ", "_")
        # defaultní filtr podle první volby
        first_ext = formats[0]
        default_name = f"{base_name}.{first_ext}"
    
        path, _ = QFileDialog.getSaveFileName(self, "Save export as…", default_name,
                                              "All files (*.*)")
        if not path:
            return
    
        base_no_ext, _sep, _ext = path.rpartition(".")
        if not base_no_ext:
            # uživatel nedal příponu → použij base_name z dialogu
            base_no_ext = path
    
        # vyber záznamy dle boards + (NOVĚ) dle scope "selected"
        records = list(getattr(self, "records", []))
        if boards_sel is not None:
            wh = set(boards_sel)
            records = [r for r in records if (getattr(r, "board", None) in wh)]
    
        if scope == "selected" and selected_file_names:
            # porovnáme podle názvu souboru (rec.path.name), jak je zobrazen v tabulce
            names = set(selected_file_names)
            def _match_selected(rec) -> bool:
                try:
                    p = getattr(rec, "path", None)
                    return bool(p and p.name in names)
                except Exception:
                    return False
            records = [r for r in records if _match_selected(r)]
    
        # připrav řádky (jen vybraná pole)
        # každý řádek → list hodnot podle (label, key) pořadí
        rows: list[list[str]] = []
        headers = [lab for (lab, _key) in fields]
    
        def _get_val(rec, key: str) -> str:
            if key == "file_name":
                try:
                    return rec.path.name if getattr(rec, "path", None) else ""
                except Exception:
                    return ""
            v = getattr(rec, key, None)
            return "" if v is None else str(v)
    
        for rec in records:
            rows.append([_get_val(rec, key) for (_lab, key) in fields])
    
        if not rows:
            QMessageBox.information(self, "Export", "No rows to export for the chosen scope/filters.")
            return
    
        # exporty
        ok, errs = [], []
    
        if "csv" in formats:
            try:
                fn = base_no_ext + ".csv"
                self._export_to_csv(fn, headers, rows)
                ok.append(fn)
            except Exception as e:
                errs.append(f"CSV: {e}")
    
        if "txt" in formats:
            try:
                fn = base_no_ext + ".txt"
                self._export_to_txt(fn, headers, rows)
                ok.append(fn)
            except Exception as e:
                errs.append(f"TXT: {e}")
    
        if "xlsx" in formats:
            try:
                fn = base_no_ext + ".xlsx"
                self._export_to_xlsx(fn, headers, rows)
                ok.append(fn)
            except ImportError:
                errs.append("XLSX: optional dependency 'openpyxl' not installed.")
            except Exception as e:
                errs.append(f"XLSX: {e}")
    
        # výsledek
        msg = []
        if ok:   msg.append("Exported:\n- " + "\n- ".join(ok))
        if errs: msg.append("\nIssues:\n- " + "\n- ".join(errs))
        from PySide6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Export", "\n".join(msg))
        try:
            self.statusBar().showMessage("Export done.")
        except Exception:
            pass

    def _export_to_csv(self, filename: str, headers: list[str], rows: list[list[str]]) -> None:
        import csv
        from pathlib import Path
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        with open(filename, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
                
    def _export_to_txt(self, path: str, headers: list[str], rows: list[list[str]]) -> None:
        """
        Rich TXT report shared by Overview & Sorted exports.
    
        Structure per record:
          Title:  Institution — Candidate — [Board]  (fallback to File name)
          Basic info:
            - Board: ...
            - Application Type: ...
            - Institution Name: ...
            - Candidate Name: ...
            - Signature Date: ...
            - File name: ...
          Sections (only if data present):
            Recognition:
              - Academia Recognition: ...
              - Certified Recognition: ...
            Contact:
              - Full Name: ...
              - Email Address: ...
              - Phone Number: ...
              - Postal Address: ...
            Curriculum:
              - Syllabi Integration: ...
              - Courses/Modules: ...
            Evidence:
              - Proof of ISTQB Certifications: ...
              - Additional Info/Documents: ...
            Links:
              - University Links: ...
        """
        def _hmap(hs: list[str]) -> dict[str, int]:
            return {h: i for i, h in enumerate(hs)}
    
        def _get(row: list[str], h2i: dict[str, int], label: str, *alts: str) -> str:
            for k in (label, *alts):
                if k in h2i:
                    idx = h2i[k]
                    if 0 <= idx < len(row):
                        val = row[idx]
                        return "" if val is None else str(val)
            return ""
    
        def _lines(val: str) -> list[str]:
            if not val:
                return []
            s = str(val).replace("\r\n", "\n").replace("\r", "\n").strip()
            if not s:
                return []
            return [ln.rstrip() for ln in s.split("\n")]
    
        def _write_bullet(fh, label: str, value: str) -> bool:
            ls = _lines(value)
            if not ls:
                return False
            # first line
            fh.write(f"- {label}: {ls[0]}\n")
            # following lines as indented bullets
            for ln in ls[1:]:
                if ln.strip():
                    fh.write(f"  {ln}\n")
            return True
    
        # Section definitions by labels (must match header labels)
        # uvnitř _export_to_txt(...) nahraď blok SECTIONS za rozšířený:
        SECTIONS: list[tuple[str, list[str]]] = [
            ("Consent", [
                "Printed Name, Title",
                "Signature Date",
            ]),
            ("For ISTQB Academia Purpose Only", [
                "Receiving Member Board",
                "Date Received",
                "Validity Start Date",
                "Validity End Date",
            ]),
            ("Recognition", [
                "Academia Recognition",
                "Certified Recognition",
            ]),
            ("Contact", [
                "Full Name",
                "Email Address",
                "Phone Number",
                "Postal Address",
            ]),
            ("Curriculum", [
                "Syllabi Integration",
                "Courses/Modules",
            ]),
            ("Evidence", [
                "Proof of ISTQB Certifications",
                "Additional Info/Documents",
            ]),
            ("Links", [
                "University Links",
            ]),
        ]

        h2i = _hmap(headers)
    
        with open(path, "w", encoding="utf-8") as fh:
            for idx, row in enumerate(rows, start=1):
                board   = _get(row, h2i, "Board")
                inst    = _get(row, h2i, "Institution Name")
                cand    = _get(row, h2i, "Candidate Name")
                app_t   = _get(row, h2i, "Application Type")
                sigdate = _get(row, h2i, "Signature Date", "Signature date", "signature_date", "sigdate")
                fname   = _get(row, h2i, "File name", "File Name", "Filename", "filename")
    
                # -------- Title --------
                parts = []
                if inst: parts.append(inst)
                if cand: parts.append(cand)
                if board: parts.append(f"[{board}]")
                title = " — ".join(p for p in parts if p) or (fname or f"Record #{idx}")
                fh.write(title + "\n")
                fh.write("=" * len(title) + "\n\n")
    
                # -------- Basic info --------
                fh.write("Basic info:\n")
                _write_bullet(fh, "Board", board)
                _write_bullet(fh, "Application Type", app_t)
                _write_bullet(fh, "Institution Name", inst)
                _write_bullet(fh, "Candidate Name", cand)
                _write_bullet(fh, "Signature Date", sigdate)
                _write_bullet(fh, "File name", fname)
                fh.write("\n")
    
                # -------- Sections --------
                for sec_title, labels in SECTIONS:
                    # include section only if at least one label has data
                    any_data = any(_get(row, h2i, lab) for lab in labels if lab in h2i)
                    if not any_data:
                        continue
                    fh.write(f"{sec_title}:\n")
                    for lab in labels:
                        if lab in h2i:
                            _write_bullet(fh, lab, _get(row, h2i, lab))
                    fh.write("\n")
    
                # Separator between records
                if idx < len(rows):
                    fh.write("-----\n\n")
                
    def _export_to_xlsx(self, filename: str, headers: list[str], rows: list[list[str]]) -> None:
        """
        XLSX export přes optional dependency 'openpyxl'.
        Bez přidávání závislostí do projektu – pokud není k dispozici, vyhodí ImportError.
        """
        from pathlib import Path
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
    
        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception as e:
            raise ImportError("openpyxl not available") from e
    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
    
        ws.append(headers)
        # tučný header
        bold = Font(bold=True)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = bold
    
        for r in rows:
            ws.append(r)
    
        # auto šířky (jednoduše dle max délky)
        for col_idx, head in enumerate(headers, start=1):
            max_len = len(head)
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 80)
    
        wb.save(filename)

    def _filter_board(self, txt: str) -> None:
        proxy = self.table.model()
        if isinstance(proxy, RecordsModel):
            proxy.set_board(txt)

    def _filter_text(self, txt: str) -> None:
        proxy = self.table.model()
        if isinstance(proxy, RecordsModel):
            proxy.set_search(txt)
            
    def _build_contacts_tab(self) -> None:
        """
        Backward-compat shim: původní jméno zavolá aktuální builder.
        """
        self._build_board_contacts_tab()

    def _build_board_contacts_tab(self) -> None:
        """
        Board Contacts tab: Board -> (Full Name, Email)
        - QStandardItemModel (3 sloupce: Board, Full Name, Email)
        - Board = needitovatelný; Full Name/Email = editovatelné
        - JSON perzistence (contacts.json), import CSV
        - auto-fit sloupců
        - NOVĚ: tlačítko Help s nápovědou a ukázkovým CSV (náhled + uložení)
        """
        from PySide6.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QWidget, QTableView, QPushButton, QMessageBox
        )
        from PySide6.QtGui import QStandardItemModel, QStandardItem
        from PySide6.QtCore import Qt, QTimer
        from PySide6.QtWidgets import QHeaderView
    
        # Najdi/usel tab widget:
        tab_widget = getattr(self, "board_contacts_tab", None)
        if tab_widget is None:
            tab_widget = getattr(self, "contacts_tab", None)
        if tab_widget is None:
            tab_widget = QWidget()
            self.board_contacts_tab = tab_widget
            try:
                self.tabs.addTab(tab_widget, "Board Contacts")
            except Exception:
                pass
        else:
            # přejmenuj tab (pokud existuje v QTabWidget)
            try:
                idx = self.tabs.indexOf(tab_widget)
                if idx >= 0:
                    self.tabs.setTabText(idx, "Board Contacts")
            except Exception:
                pass
    
        layout = QVBoxLayout(tab_widget)
    
        # Ovládací řádek
        bar = QHBoxLayout()
        self.btn_contacts_help = QPushButton("Help")
        self.btn_contacts_import = QPushButton("Import CSV…")
        self.btn_contacts_save = QPushButton("Save")
        self.btn_contacts_reload = QPushButton("Reload")
        bar.addWidget(self.btn_contacts_help)
        bar.addSpacing(8)
        bar.addWidget(self.btn_contacts_import)
        bar.addStretch(1)
        bar.addWidget(self.btn_contacts_reload)
        bar.addWidget(self.btn_contacts_save)
        layout.addLayout(bar)
    
        # Tabulka
        self.tbl_contacts = QTableView(tab_widget)
        self.tbl_contacts.setSelectionBehavior(QTableView.SelectRows)
        self.tbl_contacts.setSelectionMode(QTableView.ExtendedSelection)
        self.tbl_contacts.setSortingEnabled(True)
    
        # Model
        self._contacts_headers = ["Board", "Full Name", "Email"]
        self._contacts_model = QStandardItemModel(0, len(self._contacts_headers), self)
        self._contacts_model.setHorizontalHeaderLabels(self._contacts_headers)
        self.tbl_contacts.setModel(self._contacts_model)
    
        # Fitting sloupců
        hdr = self.tbl_contacts.horizontalHeader()
        hdr.setStretchLastSection(True)
        try:
            hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
            hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
            hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        except Exception:
            hdr.setResizeMode(0, QHeaderView.ResizeToContents)
            hdr.setResizeMode(1, QHeaderView.ResizeToContents)
            hdr.setResizeMode(2, QHeaderView.ResizeToContents)
    
        layout.addWidget(self.tbl_contacts, 1)
    
        # Data -> model
        self._contacts_rebuild_model()
    
        # Akce
        self.btn_contacts_save.clicked.connect(lambda: self._save_contacts_json(self._contacts_collect_data()))
        self.btn_contacts_reload.clicked.connect(self._contacts_rebuild_model)
        self.btn_contacts_import.clicked.connect(self._contacts_import_csv)
        self.btn_contacts_help.clicked.connect(self._contacts_show_help)
    
        # Finální doladění velikostí po vykreslení
        def _refit():
            try:
                for c in range(self._contacts_model.columnCount()):
                    self.tbl_contacts.resizeColumnToContents(c)
            except Exception:
                pass
        QTimer.singleShot(0, _refit)
        
    def _contacts_show_help(self) -> None:
        """
        Zobrazí dialog s nápovědou/importními tipy a ukázkovým CSV.
        Umožní uložit šablonu CSV do souboru.
        """
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QLabel, QPlainTextEdit,
            QDialogButtonBox, QPushButton, QFileDialog, QWidget, QHBoxLayout
        )
        from PySide6.QtCore import Qt
    
        dlg = QDialog(self)
        dlg.setWindowTitle("Board Contacts — Help")
        lay = QVBoxLayout(dlg)
    
        help_text = (
            "CSV columns (case-insensitive):\n"
            "  - board\n"
            "  - full_name (or 'Full Name' / 'Name')\n"
            "  - email (or 'E-mail')\n\n"
            "Rows are matched by 'board'. Unknown boards are ignored (not added).\n"
            "Empty values are allowed.\n\n"
            "Example CSV:\n"
        )
        lay.addWidget(QLabel(help_text))
    
        sample = self._contacts_sample_csv()
    
        viewer = QPlainTextEdit()
        viewer.setReadOnly(True)
        viewer.setPlainText(sample)
        viewer.setLineWrapMode(QPlainTextEdit.NoWrap)
        lay.addWidget(viewer, 1)
    
        # Buttons: Save template… + Close
        btns_bar = QHBoxLayout()
        btn_save = QPushButton("Save CSV template…")
        btns_bar.addWidget(btn_save)
        btns_bar.addStretch(1)
        lay.addLayout(btns_bar)
    
        def _save():
            path, _ = QFileDialog.getSaveFileName(dlg, "Save CSV template…", "board_contacts_template.csv", "CSV Files (*.csv);;All Files (*)")
            if not path:
                return
            try:
                with open(path, "w", encoding="utf-8") as fh:
                    fh.write(sample)
            except Exception as e:
                # Plain info — nenaruší běh
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(dlg, "Save CSV template", f"Failed to save:\n{e}")
    
        btn_save.clicked.connect(_save)
    
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dlg.reject)
        buttons.accepted.connect(dlg.accept)
        lay.addWidget(buttons)
    
        dlg.resize(720, 420)
        dlg.exec_()
        
    def _contacts_sample_csv(self) -> str:
        """
        Vrátí ukázkové CSV s hlavičkou a několika boardy.
        """
        try:
            boards = sorted(KNOWN_BOARDS)
        except Exception:
            boards = ["ATB", "CSTB", "ISTQB"]
        # Vezmi prvních pár pro příklad
        sample_boards = boards[:3] if len(boards) >= 3 else boards
        lines = ["board,full_name,email"]
        for b in sample_boards:
            email = f"{''.join(ch for ch in b.lower() if ch.isalnum())}-liaison@example.org"
            lines.append(f"{b},Contact for {b},{email}")
        return "\n".join(lines) + "\n"
        
    def _contacts_json_path(self):
        """
        Umístění JSONu s kontakty v kořeni repozitáře (vedle zdrojáků).
        Vhodné pro .gitignore.
        """
        from pathlib import Path
        # repo_root ≈ dva levely nad tímto souborem: app/main_window.py -> repo/
        repo_root = Path(__file__).resolve().parents[2]
        return repo_root / "contacts.json"
    
    def _load_contacts_json(self) -> dict:
        """
        Načti JSON kontaktů: { "<BOARD>": { "full_name": str, "email": str }, ... }
        Neexistuje-li soubor, vrať {}.
        """
        import json
        p = self._contacts_json_path()
        try:
            if p.exists():
                with p.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    if isinstance(data, dict):
                        return data
        except Exception:
            pass
        return {}
    
    def _save_contacts_json(self, data: dict) -> None:
        """
        Ulož JSON kontaktů. Vytvoří/aktualizuje contacts.json.
        """
        import json
        from pathlib import Path
        p = self._contacts_json_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            with p.open("w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False, indent=2)
            try:
                self.statusBar().showMessage("Contacts saved.")
            except Exception:
                pass
        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Save contacts", f"Failed to save contacts.json:\n{e}")
            
    def _contacts_rebuild_model(self) -> None:
        """
        Naplní model kontakty pro všechny KNOWN_BOARDS.
        Výchozí hodnoty prázdné, pokud nejsou v JSON.
        """
        from PySide6.QtGui import QStandardItem
        from PySide6.QtCore import Qt
    
        data = self._load_contacts_json()  # dict
        self._contacts_model.setRowCount(0)
    
        # Vždy zobraz všechny boardy; jméno/email mohou být prázdné
        boards = sorted(KNOWN_BOARDS)
        for b in boards:
            full = ""
            mail = ""
            try:
                rec = data.get(b, {})
                full = rec.get("full_name", "") or ""
                mail = rec.get("email", "") or ""
            except Exception:
                pass
    
            it_board = QStandardItem(b)
            it_board.setEditable(False)
            it_board.setData(b, Qt.DisplayRole)
    
            it_full = QStandardItem(full)
            it_full.setEditable(True)
    
            it_mail = QStandardItem(mail)
            it_mail.setEditable(True)
    
            self._contacts_model.appendRow([it_board, it_full, it_mail])
    
        # sloupce přizpůsob po naplnění
        try:
            view = self.tbl_contacts
            for c in range(self._contacts_model.columnCount()):
                view.resizeColumnToContents(c)
        except Exception:
            pass
        
    def _contacts_collect_data(self) -> dict:
        """
        Čti aktuální hodnoty z modelu a vytvoř JSON strukturu:
        { "BOARD": {"full_name": "...", "email": "..."}, ... }
        Prázdné dvojice ukládám také (výslovně prázdné).
        """
        from PySide6.QtCore import Qt
        out: dict = {}
        rows = self._contacts_model.rowCount()
        for r in range(rows):
            board = self._contacts_model.index(r, 0).data(Qt.DisplayRole) or ""
            full  = self._contacts_model.index(r, 1).data(Qt.DisplayRole) or ""
            mail  = self._contacts_model.index(r, 2).data(Qt.DisplayRole) or ""
            if board:
                out[board] = {"full_name": str(full), "email": str(mail)}
        return out
    
    def _contacts_import_csv(self) -> None:
        """
        Import CSV s hlavičkami (case-insensitive; varianty akceptované):
          - board | Board
          - full_name | Full Name | Name
          - email | Email | E-mail
        Řádky se mapují podle 'board'. Nematchnuté boardy se ignorují (nepřidáváme nové).
        """
        import csv
        from PySide6.QtWidgets import QFileDialog, QMessageBox
    
        path, _ = QFileDialog.getOpenFileName(self, "Import contacts (CSV)", "", "CSV Files (*.csv);;All Files (*)")
        if not path:
            return
    
        # přečti CSV
        rows = []
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as fh:
                rdr = csv.DictReader(fh)
                # normalizace klíčů
                norm = lambda s: (s or "").strip().lower()
                for rec in rdr:
                    if not isinstance(rec, dict):
                        continue
                    row = {norm(k): (v or "").strip() for k, v in rec.items()}
                    rows.append(row)
        except Exception as e:
            QMessageBox.warning(self, "Import CSV", f"Failed to read CSV:\n{e}")
            return
    
        if not rows:
            QMessageBox.information(self, "Import CSV", "No rows found in the CSV file.")
            return
    
        # mapování názvů sloupců
        def pick(d: dict, keys: list[str]) -> str:
            for k in keys:
                if k in d and d[k]:
                    return d[k]
            return ""
    
        # Pro rychlý update – vytvoř mapu board -> (full, mail)
        updates: dict[str, tuple[str, str]] = {}
        for d in rows:
            b = pick(d, ["board"])
            if not b:
                continue
            full = pick(d, ["full_name", "full name", "name"])
            mail = pick(d, ["email", "e-mail", "mail"])
            updates[b] = (full, mail)
    
        if not updates:
            QMessageBox.information(self, "Import CSV", "No usable data (missing 'board' column).")
            return
    
        # aplikuj do modelu – jen u existujících boardů
        from PySide6.QtCore import Qt
        changed = 0
        for r in range(self._contacts_model.rowCount()):
            board = self._contacts_model.index(r, 0).data(Qt.DisplayRole)
            if board in updates:
                full, mail = updates[board]
                if full:
                    self._contacts_model.setData(self._contacts_model.index(r, 1), full, Qt.EditRole)
                if mail:
                    self._contacts_model.setData(self._contacts_model.index(r, 2), mail, Qt.EditRole)
                changed += 1
    
        QMessageBox.information(self, "Import CSV", f"Imported/updated contacts for {changed} board(s).")

    def _build_sorted_tab(self) -> None:
        """
        Záložka 'Sorted PDFs': vlevo strom, vpravo formulář.
        Přidány editory pro: Printed Name, Title; Receiving Member Board; Date Received; Validity Start/End.
        Navíc přidán řádek 'File name' (read-only), aby nepadalo _sorted_fill_details().
        """
        from PySide6.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QTreeWidget, QTreeWidgetItem, QSplitter,
            QWidget, QFormLayout, QLineEdit, QPlainTextEdit, QPushButton, QSizePolicy
        )
        from PySide6.QtCore import Qt, QTimer
    
        layout = QVBoxLayout()
        self.split_sorted = QSplitter(self.sorted_tab)
        self.split_sorted.setOrientation(Qt.Horizontal)
    
        # LEVÁ strana – strom
        self.tree_sorted = QTreeWidget()
        self.tree_sorted.setHeaderLabels(["Board / PDF"])
        self.tree_sorted.itemSelectionChanged.connect(self._sorted_on_item_changed)
    
        # PRAVÁ strana – formulář
        right = QWidget()
        right_layout = QVBoxLayout(right)
        self.form_sorted = QFormLayout()
        self.form_sorted.setRowWrapPolicy(QFormLayout.WrapLongRows)
        self.form_sorted.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
    
        def _mkline() -> QLineEdit:
            e = QLineEdit()
            e.setClearButtonEnabled(True)
            return e
        def _mktxt() -> QPlainTextEdit:
            t = QPlainTextEdit()
            t.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            t.setMaximumBlockCount(100000)
            return t
    
        # Základní pole
        self.ed_board = _mkline(); self.ed_board.setReadOnly(True)
        self.ed_app_type = _mkline()
        self.ed_inst_name = _mkline()
        self.ed_cand_name = _mkline()
        self.ed_rec_acad = _mkline()
        self.ed_rec_cert = _mkline()
        self.ed_fullname = _mkline()
        self.ed_email = _mkline()
        self.ed_phone = _mkline()
        self.ed_address = _mktxt()
        self.ed_syllabi = _mktxt()
        self.ed_courses = _mktxt()
        self.ed_proof = _mktxt()
        self.ed_links = _mktxt()
        self.ed_additional = _mktxt()
        # Consent
        self.ed_printed_name_title = _mkline()   # NOVĚ
        self.ed_sigdate = _mkline()
        # ISTQB internal
        self.ed_receiving_member_board = _mkline()  # NOVĚ
        self.ed_date_received = _mkline()           # NOVĚ
        self.ed_validity_start_date = _mkline()     # NOVĚ
        self.ed_validity_end_date = _mkline()       # NOVĚ
        # File name (read-only, aby se necrashovalo)
        self.ed_filename = _mkline(); self.ed_filename.setReadOnly(True)
    
        # Alias (stávající jména používaná jinde v kódu nechávám)
        self.ed_contact_full_name = self.ed_fullname
        self.ed_contact_email = self.ed_email
        self.ed_contact_phone = self.ed_phone
        self.ed_postal_address = self.ed_address
        self.ed_syllabi_integration_description = self.ed_syllabi
        self.ed_courses_modules_list = self.ed_courses
        self.ed_proof_of_istqb_certifications = self.ed_proof
        self.ed_university_links = self.ed_links
        self.ed_additional_information_documents = self.ed_additional
    
        # Form – pořadí (Printed Name před Signature Date; sekce 7 za ním)
        self.form_sorted.addRow("Board:", self.ed_board)
        self.form_sorted.addRow("Application Type:", self.ed_app_type)
        self.form_sorted.addRow("Institution Name:", self.ed_inst_name)
        self.form_sorted.addRow("Candidate Name:", self.ed_cand_name)
        self.form_sorted.addRow("Academia Recognition:", self.ed_rec_acad)
        self.form_sorted.addRow("Certified Recognition:", self.ed_rec_cert)
        self.form_sorted.addRow("Full Name:", self.ed_fullname)
        self.form_sorted.addRow("Email Address:", self.ed_email)
        self.form_sorted.addRow("Phone Number:", self.ed_phone)
        self.form_sorted.addRow("Postal Address:", self.ed_address)
        self.form_sorted.addRow("Syllabi Integration:", self.ed_syllabi)
        self.form_sorted.addRow("Courses/Modules:", self.ed_courses)
        self.form_sorted.addRow("Proof of ISTQB Certifications:", self.ed_proof)
        self.form_sorted.addRow("University Links:", self.ed_links)
        self.form_sorted.addRow("Additional Info/Documents:", self.ed_additional)
        # Consent
        self.form_sorted.addRow("Printed Name, Title:", self.ed_printed_name_title)
        self.form_sorted.addRow("Signature Date:", self.ed_sigdate)
        # ISTQB internal
        self.form_sorted.addRow("Receiving Member Board:", self.ed_receiving_member_board)
        self.form_sorted.addRow("Date Received:", self.ed_date_received)
        self.form_sorted.addRow("Validity Start Date:", self.ed_validity_start_date)
        self.form_sorted.addRow("Validity End Date:", self.ed_validity_end_date)
        # File name
        self.form_sorted.addRow("File name:", self.ed_filename)
    
        right_layout.addLayout(self.form_sorted)
    
        # Tlačítka
        btn_row = QHBoxLayout()
        self.btn_sorted_edit = QPushButton("Edit")
        self.btn_sorted_save = QPushButton("Save to DB")
        self.btn_sorted_export = QPushButton("Export…")
        self.btn_sorted_rescan = QPushButton("Rescan Sorted")
        btn_row.addWidget(self.btn_sorted_edit)
        btn_row.addWidget(self.btn_sorted_save)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_sorted_export)
        btn_row.addWidget(self.btn_sorted_rescan)
        right_layout.addLayout(btn_row)
    
        self.btn_sorted_edit.clicked.connect(lambda: self._sorted_set_editable(True))
        self.btn_sorted_save.clicked.connect(self._sorted_save_changes)
        self.btn_sorted_export.clicked.connect(self.export_sorted_db)
        self.btn_sorted_rescan.clicked.connect(self.rescan_sorted)
    
        # Výchozí – read-only; Board/File name vždy read-only
        self._sorted_set_editable(False)
    
        # Osazení splitteru a layoutu
        self.split_sorted.addWidget(self.tree_sorted)
        self.split_sorted.addWidget(right)
        self.split_sorted.setStretchFactor(0, 2)
        self.split_sorted.setStretchFactor(1, 1)
        layout.addWidget(self.split_sorted, 1)
        self.sorted_tab.setLayout(layout)
        
    def export_sorted_db(self) -> None:
        """
        Export CURRENT dataset from 'Sorted PDFs' tab using data stored in DB (self.sorted_db),
        s nezávislostí na SortedDb.boards().
        """
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QCheckBox,
            QListWidget, QListWidgetItem, QDialogButtonBox, QFileDialog, QMessageBox
        )
        from PySide6.QtCore import Qt
        from datetime import datetime
        import os
        from pathlib import Path
    
        # Columns (labels/keys) – stejné jako v Overview
        FIELDS: list[tuple[str, str]] = [
            ("Board", "board"),
            ("Application Type", "application_type"),
            ("Institution Name", "institution_name"),
            ("Candidate Name", "candidate_name"),
            ("Academia Recognition", "recognition_academia"),
            ("Certified Recognition", "recognition_certified"),
            ("Full Name", "contact_full_name"),
            ("Email Address", "contact_email"),
            ("Phone Number", "contact_phone"),
            ("Postal Address", "contact_postal_address"),
            ("Syllabi Integration", "syllabi_integration_description"),
            ("Courses/Modules", "courses_modules_list"),
            ("Proof of ISTQB Certifications", "proof_of_istqb_certifications"),
            ("University Links", "university_links"),
            ("Additional Info/Documents", "additional_information_documents"),
            ("Printed Name, Title", "printed_name_title"),
            ("Signature Date", "signature_date"),
            ("Receiving Member Board", "receiving_member_board"),
            ("Date Received", "date_received"),
            ("Validity Start Date", "validity_start_date"),
            ("Validity End Date", "validity_end_date"),
            ("File name", "file_name"),
        ]
    
        # Pomocná: iterace DB položek (nezávisle na boards())
        def _iter_sorted_items():
            it = getattr(self.sorted_db, "iter_items", None)
            if callable(it):
                yield from it()
                return
            try:
                for k, v in self.sorted_db.items():
                    yield Path(k), v
            except Exception:
                return
    
        # Zjisti dostupné boardy z dat
        boards_avail = set()
        for abs_path, rec in _iter_sorted_items():
            try:
                data = rec.get("data", {}) if isinstance(rec, dict) else {}
                board = data.get("board")
                if board:
                    boards_avail.add(board)
            except Exception:
                continue
        boards_avail = sorted(boards_avail)
    
        # Dialog – jednoduchá verze (formáty + výběr boards + výběr polí)
        class ExportDialog(QDialog):
            def __init__(self, boards_avail: list[str], parent=None):
                super().__init__(parent)
                self.setWindowTitle("Export options")
                main = QVBoxLayout(self)
    
                gb_formats = QGroupBox("Formats")
                lay_f = QHBoxLayout(gb_formats)
                self.cb_xlsx = QCheckBox("XLSX"); self.cb_xlsx.setChecked(True)
                self.cb_csv  = QCheckBox("CSV")
                self.cb_txt  = QCheckBox("TXT")
                lay_f.addWidget(self.cb_xlsx); lay_f.addWidget(self.cb_csv); lay_f.addWidget(self.cb_txt)
                main.addWidget(gb_formats)
    
                gb_board = QGroupBox("Boards")
                lay_b = QVBoxLayout(gb_board)
                self.list_boards = QListWidget()
                self.list_boards.setSelectionMode(QListWidget.MultiSelection)
                for b in boards_avail:
                    it = QListWidgetItem(b)
                    it.setSelected(True)  # defaultně všechny
                    self.list_boards.addItem(it)
                lay_b.addWidget(self.list_boards)
                main.addWidget(gb_board)
    
                gb_fields = QGroupBox("Fields (order)")
                lay_fi = QVBoxLayout(gb_fields)
                self.list_fields = QListWidget()
                self.list_fields.setSelectionMode(QListWidget.MultiSelection)
                for label, key in FIELDS:
                    it = QListWidgetItem(label)
                    it.setData(Qt.UserRole, key)
                    it.setSelected(True)
                    self.list_fields.addItem(it)
                lay_fi.addWidget(self.list_fields)
                main.addWidget(gb_fields)
    
                btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btns.accepted.connect(self.accept)
                btns.rejected.connect(self.reject)
                main.addWidget(btns)
    
        dlg = ExportDialog(boards_avail, self)
        if dlg.exec() != QDialog.Accepted:
            return
    
        fmt_xlsx = dlg.cb_xlsx.isChecked()
        fmt_csv  = dlg.cb_csv.isChecked()
        fmt_txt  = dlg.cb_txt.isChecked()
        selected_fields = [(dlg.list_fields.item(i).text(), dlg.list_fields.item(i).data(Qt.UserRole))
                           for i in range(dlg.list_fields.count())
                           if dlg.list_fields.item(i).isSelected()]
        if not selected_fields:
            QMessageBox.warning(self, "Export", "No fields selected.")
            return
    
        boards_sel = {dlg.list_boards.item(i).text()
                      for i in range(dlg.list_boards.count())
                      if dlg.list_boards.item(i).isSelected()}
        if not boards_sel:
            boards_sel = set(boards_avail)
    
        headers = [label for (label, _) in selected_fields]
        rows = []
        for abs_path, rec in _iter_sorted_items():
            try:
                data = rec.get("data", {}) if isinstance(rec, dict) else {}
                b = data.get("board")
                if b not in boards_sel:
                    continue
                row = [str(data.get(k, "") or "") for (_, k) in selected_fields]
                rows.append(row)
            except Exception:
                continue
    
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"sorted_export_{now}"
        dlg_dir = os.getcwd()
        path, _ = QFileDialog.getSaveFileName(self, "Save export", os.path.join(dlg_dir, filename_base + ".xlsx"),
                                              "XLSX (*.xlsx);;CSV (*.csv);;TXT (*.txt)")
        if not path:
            return
    
        if fmt_xlsx:
            self._export_to_xlsx(path if path.lower().endswith(".xlsx") else path + ".xlsx", headers, rows)
        if fmt_csv:
            self._export_to_csv(path if path.lower().endswith(".csv") else path + ".csv", headers, rows)
        if fmt_txt:
            self._export_to_txt(path if path.lower().endswith(".txt") else path + ".txt", headers, rows)
    
        QMessageBox.information(self, "Export", "Export finished.")
        
    def rescan_sorted(self) -> None:
        """
        Sestaví strom 'Sorted PDFs' z DB (edited i parsed).
        Do UserRole ukládá DB klíč (relativní cesta v rámci sorted_root) a také absolutní
        cestu v sorted_root pro případ exportů. Nepoužívá SortedDb.boards().
        """
        from PySide6.QtCore import Qt
        from PySide6.QtWidgets import QTreeWidgetItem
        from pathlib import Path
    
        self.tree_sorted.clear()
    
        sorted_root = getattr(self.sorted_db, "sorted_root", None)
        sorted_root = Path(sorted_root).resolve() if sorted_root else None
    
        # Iterace položek z DB
        def _iter_sorted_items():
            it = getattr(self.sorted_db, "iter_items", None)
            if callable(it):
                # očekáváme, že iter_items() vrací (abs_sorted_path, record)
                yield from it()
                return
            # fallback: pokud SortedDb nemá iter_items a má .items() s klíči = relativním klíčem
            try:
                for rel_key, rec in self.sorted_db.items():
                    abs_sorted = (sorted_root / rel_key).resolve() if sorted_root else Path(rel_key).resolve()
                    yield abs_sorted, rec
            except Exception:
                return
    
        root_items: dict[str, QTreeWidgetItem] = {}
    
        def _ensure_board_node(board_name: str) -> QTreeWidgetItem:
            node = root_items.get(board_name)
            if node is None:
                node = QTreeWidgetItem([board_name])
                node.setData(0, Qt.UserRole, None)
                self.tree_sorted.addTopLevelItem(node)
                root_items[board_name] = node
            return node
    
        for abs_sorted_path, rec in _iter_sorted_items():
            try:
                if isinstance(abs_sorted_path, str):
                    abs_sorted_path = Path(abs_sorted_path)
                data = rec.get("data", {}) if isinstance(rec, dict) else {}
                board = data.get("board") or abs_sorted_path.parent.name
                name = data.get("file_name") or abs_sorted_path.name
    
                # DB klíč = cesta relativně k sorted_root, pokud existuje
                db_key = None
                if sorted_root:
                    try:
                        db_key = str(abs_sorted_path.resolve().relative_to(sorted_root))
                    except Exception:
                        db_key = None
    
                parent = _ensure_board_node(board)
                child = QTreeWidgetItem([name])
                # UserRole: uložíme DB klíč (který očekává SortedDb.get/key_for)
                # a navíc ABS path v sorted_root (pro případné exporty).
                child.setData(0, Qt.UserRole, db_key or "")
                child.setData(0, Qt.UserRole + 1, str(abs_sorted_path.resolve()))
                parent.addChild(child)
            except Exception:
                continue
    
        self.tree_sorted.expandAll()

    def _sorted_db_path(self) -> Path:
        return self.sorted_db.db_path
    
    def _sorted_key_for(self, abs_path: Path) -> str:
        return self.sorted_db.key_for(abs_path)
    
    def _sorted_on_item_changed(self) -> None:
        """Výběr ve stromu 'Sorted PDFs' → ulož kontext a naplň detail."""
        from pathlib import Path
        from PySide6.QtCore import Qt
    
        sel = self.tree_sorted.selectedItems()
        if not sel:
            self._sorted_fill_details(None)
            return
    
        item = sel[0]
        parent = item.parent()
        board = parent.text(0) if parent else ""
        file_name = item.text(0)
    
        db_key = item.data(0, Qt.UserRole)            # relativní klíč vůči sorted_root
        abs_sorted = item.data(0, Qt.UserRole + 1)    # absolutní cesta v 'Sorted PDFs'
    
        # ulož pro _sorted_fill_details
        self._sorted_sel_db_key = db_key or ""
        self._sorted_sel_abs_sorted = abs_sorted or ""
        self._sorted_sel_board = board or ""
        self._sorted_sel_file_name = file_name or ""
    
        # preferuj ABS cestu v sorted_root, ale vlastní naplnění si ještě dopočítá kandidátní cesty
        self._sorted_fill_details(Path(abs_sorted) if abs_sorted else None)
    
    def _sorted_fill_details(self, abs_path: Optional[Path]) -> None:
        from pathlib import Path
    
        def _set(w, val: str):
            if hasattr(w, "setPlainText"):
                w.setPlainText(val or "")
            else:
                w.setText(val or "")
    
        def _blank():
            for w in (self.ed_board, self.ed_app_type, self.ed_inst_name, self.ed_cand_name,
                      self.ed_rec_acad, self.ed_rec_cert, self.ed_fullname, self.ed_email,
                      self.ed_phone, self.ed_address, self.ed_syllabi, self.ed_courses,
                      self.ed_proof, self.ed_links, self.ed_additional,
                      self.ed_printed_name_title, self.ed_sigdate,
                      self.ed_receiving_member_board, self.ed_date_received,
                      self.ed_validity_start_date, self.ed_validity_end_date):
                _set(w, "")
            _set(self.ed_filename, "")
            self._sorted_set_editable(False)
            self._sorted_set_status(None)
    
        if abs_path is None:
            # zkusí si ještě složit kandidáty z uloženého kontextu
            pass
    
        # 1) načti DB, pokud to jde (pouze pokud je cesta v sorted_root)
        rec = None
        try:
            if abs_path is not None:
                rec = self.sorted_db.get(abs_path)  # může vyžadovat path pod sorted_root
        except Exception:
            rec = None
    
        data = rec.get("data", {}) if rec else {}
    
        # 2) sestav KANDIDÁTNÍ cesty k PDF v pořadí:
        #    a) původní cesta uložená v DB (data["path"])
        #    b) absolutní cesta v sorted_root (UserRole+1 / abs_path)
        #    c) odvozená cesta: pdf_root / <board> / <file_name>
        candidates: list[Path] = []
        if isinstance(data, dict):
            p = data.get("path")
            if p:
                try:
                    pp = Path(p)
                    candidates.append(pp)
                except Exception:
                    pass
    
        if abs_path:
            candidates.append(abs_path)
    
        try:
            board = getattr(self, "_sorted_sel_board", "") or ""
            fname = getattr(self, "_sorted_sel_file_name", "") or ""
            if getattr(self, "pdf_root", None) and board and fname:
                p3 = Path(self.pdf_root) / board / fname
                candidates.append(p3)
        except Exception:
            pass
    
        # 3) pokud je DB nekompletní nebo prázdná, zkus parse prvního existujícího kandidáta
        REQUIRED_KEYS = [
            "board", "application_type", "institution_name", "candidate_name",
            "recognition_academia", "recognition_certified",
            "contact_full_name", "contact_email", "contact_phone", "contact_postal_address",
            "syllabi_integration_description", "courses_modules_list",
            "proof_of_istqb_certifications", "university_links", "additional_information_documents",
            "printed_name_title", "signature_date",
            "receiving_member_board", "date_received",
            "validity_start_date", "validity_end_date",
            "file_name",
        ]
    
        def _is_incomplete(d: dict) -> bool:
            if not d:
                return True
            for k in REQUIRED_KEYS:
                if k not in d:  # klíč chybí → doplnit z parsingu
                    return True
            return False
    
        merged = dict(data) if isinstance(data, dict) else {}
    
        if _is_incomplete(merged):
            try:
                from app.pdf_scanner import PdfScanner
                from dataclasses import asdict
                # projdi kandidáty, první existující zparsuj a doplň chybějící klíče
                for cand in candidates:
                    try:
                        if not cand:
                            continue
                        cpath = Path(cand).resolve()
                        if not cpath.exists():
                            continue
                        for r in PdfScanner(cpath.parent).scan():
                            try:
                                rp = Path(r.path).resolve() if getattr(r, "path", None) else None
                            except Exception:
                                rp = None
                            if rp and rp == cpath:
                                parsed = asdict(r)
                                # slouč: existující DB hodnoty nechme, doplňme jen chybějící
                                for k, v in parsed.items():
                                    if (k not in merged) or (merged.get(k) in (None, "")):
                                        merged[k] = v
                                break
                        if not _is_incomplete(merged):
                            break
                    except Exception:
                        continue
            except Exception:
                pass  # když parse selže, zobrazíme co je
    
        # 4) Naplň UI — prázdné zůstává prázdné
        def get(k: str) -> str:
            return "" if not merged else str(merged.get(k, "") or "")
    
        _set(self.ed_board, get("board"))
        _set(self.ed_app_type, get("application_type"))
        _set(self.ed_inst_name, get("institution_name"))
        _set(self.ed_cand_name, get("candidate_name"))
        _set(self.ed_rec_acad, get("recognition_academia"))
        _set(self.ed_rec_cert, get("recognition_certified"))
        _set(self.ed_fullname, get("contact_full_name"))
        _set(self.ed_email, get("contact_email"))
        _set(self.ed_phone, get("contact_phone"))
        _set(self.ed_address, get("contact_postal_address"))
        _set(self.ed_syllabi, get("syllabi_integration_description"))
        _set(self.ed_courses, get("courses_modules_list"))
        _set(self.ed_proof, get("proof_of_istqb_certifications"))
        _set(self.ed_links, get("university_links"))
        _set(self.ed_additional, get("additional_information_documents"))
        _set(self.ed_printed_name_title, get("printed_name_title"))
        _set(self.ed_sigdate, get("signature_date"))
        _set(self.ed_receiving_member_board, get("receiving_member_board"))
        _set(self.ed_date_received, get("date_received"))
        _set(self.ed_validity_start_date, get("validity_start_date"))
        _set(self.ed_validity_end_date, get("validity_end_date"))
        # File name: vždy smysluplná hodnota
        fname = merged.get("file_name") if isinstance(merged, dict) else None
        if not fname:
            try:
                fname = getattr(self, "_sorted_sel_file_name", "") or (abs_path.name if abs_path else "")
            except Exception:
                fname = ""
        _set(self.ed_filename, fname)
    
        self._sorted_set_editable(False)
        self._sorted_set_status("Edited" if (rec and rec.get("edited")) else "Parsed (unmodified)")
        self._sorted_current_path = abs_path if abs_path else None
    
    def _sorted_set_status(self, txt: Optional[str]) -> None:
        self.lbl_sorted_status.setText("—" if not txt else txt)
    
    def _sorted_set_editable(self, can_edit: bool) -> None:
        # Board a File name necháme read-only; ostatní podle can_edit
        ro_lines = (self.ed_board, self.ed_filename)
        for w in ro_lines:
            w.setReadOnly(True)
    
        lines = (self.ed_app_type, self.ed_inst_name, self.ed_cand_name,
                 self.ed_rec_acad, self.ed_rec_cert, self.ed_fullname,
                 self.ed_email, self.ed_phone,
                 self.ed_printed_name_title, self.ed_sigdate,
                 self.ed_receiving_member_board, self.ed_date_received,
                 self.ed_validity_start_date, self.ed_validity_end_date)
        for w in lines:
            w.setReadOnly(not can_edit)
    
        texts = (self.ed_address, self.ed_syllabi, self.ed_courses,
                 self.ed_proof, self.ed_links, self.ed_additional)
        for w in texts:
            w.setReadOnly(not can_edit)

    def _sorted_save_changes(self) -> None:
        from PySide6.QtWidgets import QMessageBox
        from pathlib import Path
    
        abs_path = getattr(self, "_sorted_current_path", None)
        if not abs_path:
            QMessageBox.information(self, "Save to DB", "No PDF selected.")
            return
    
        def _get(w):
            return w.toPlainText() if hasattr(w, "toPlainText") else w.text()
    
        file_name_value = None
        if hasattr(self, "ed_filename"):
            try:
                file_name_value = _get(self.ed_filename).strip()
            except Exception:
                file_name_value = None
        if not file_name_value:
            file_name_value = Path(abs_path).name  # fallback, aby to nikdy nepadalo
    
        new_data = {
            "board": self.ed_board.text().strip(),
            "application_type": _get(self.ed_app_type).strip(),
            "institution_name": _get(self.ed_inst_name).strip(),
            "candidate_name": _get(self.ed_cand_name).strip(),
            "recognition_academia": _get(self.ed_rec_acad).strip(),
            "recognition_certified": _get(self.ed_rec_cert).strip(),
            "contact_full_name": _get(self.ed_fullname).strip(),
            "contact_email": _get(self.ed_email).strip(),
            "contact_phone": _get(self.ed_phone).strip(),
            "contact_postal_address": _get(self.ed_address).strip(),
            "syllabi_integration_description": _get(self.ed_syllabi).strip(),
            "courses_modules_list": _get(self.ed_courses).strip(),
            "proof_of_istqb_certifications": _get(self.ed_proof).strip(),
            "university_links": _get(self.ed_links).strip(),
            "additional_information_documents": _get(self.ed_additional).strip(),
            # NOVÁ POLE
            "printed_name_title": _get(self.ed_printed_name_title).strip(),
            "signature_date": _get(self.ed_sigdate).strip(),
            "receiving_member_board": _get(self.ed_receiving_member_board).strip(),
            "date_received": _get(self.ed_date_received).strip(),
            "validity_start_date": _get(self.ed_validity_start_date).strip(),
            "validity_end_date": _get(self.ed_validity_end_date).strip(),
            # meta
            "file_name": file_name_value,
            "path": str(Path(abs_path).resolve()),
        }
    
        self.sorted_db.mark_edited(Path(abs_path), new_data)
        self.sorted_db.save()
    
        self._sorted_set_editable(False)
        self._sorted_set_status("Edited")
        try:
            self.statusBar().showMessage("Saved to DB.")
        except Exception:
            pass
        
    def _filter_board_sorted(self, txt: str) -> None:
        proxy = self.table_sorted.model()
        if isinstance(proxy, RecordsModel):
            proxy.set_board(txt)
            
    def _filter_text_sorted(self, txt: str) -> None:
        proxy = self.table_sorted.model()
        if isinstance(proxy, RecordsModel):
            proxy.set_search(txt)

    def _selected_sorted_record(self) -> Optional[PdfRecord]:
        sel = self.table_sorted.selectionModel()
        if not sel or not sel.hasSelection():
            return None
        index = sel.selectedRows()[0]
        FILE_COL = 16
        proxy = self.table_sorted.model()
        if isinstance(proxy, QSortFilterProxyModel):
            sidx = proxy.mapToSource(proxy.index(index.row(), FILE_COL))
            src = proxy.sourceModel()
        else:
            sidx = index
            src = self.table_sorted.model()
        path_str = src.index(sidx.row(), FILE_COL).data(Qt.UserRole + 1)
        for r in self.records_sorted:
            if str(r.path) == path_str:
                return r
        return None
        
    from typing import Optional
    from PySide6.QtWidgets import QMessageBox
    from PySide6.QtGui import QDesktopServices
    from PySide6.QtCore import QUrl
    
    def open_selected_pdf(self) -> None:
        """Open the PDF file of the currently selected row in Overview."""
        rec = self._selected_record()
        if not rec:
            QMessageBox.information(self, "Open PDF", "Please select a row first.")
            return
        # macOS-safe: QDesktopServices.openUrl requires QUrl, not a string
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(rec.path)))

    # ----- Browser tab -----
    def _build_browser_tab(self) -> None:
        """
        PDF Browser: TOP = strom, BOTTOM = detail (formulář s lbl_*).
        Přidány položky:
          - Printed Name, Title (před Signature Date)
          - Receiving Member Board, Date Received, Validity Start/End (za Signature Date)
        """
        from PySide6.QtCore import Qt, QTimer
        from PySide6.QtWidgets import (
            QSplitter, QWidget, QVBoxLayout, QTreeView, QFormLayout, QLabel,
            QScrollArea, QFileSystemModel, QSizePolicy, QHeaderView
        )
    
        vsplit = QSplitter(Qt.Vertical, self.browser_tab)
    
        # Nahoře strom
        top_widget = QWidget(vsplit)
        top_layout = QVBoxLayout(top_widget)
    
        self.fs_model = QFileSystemModel(self)
        self.fs_model.setRootPath(str(self.pdf_root))
        self.tree = QTreeView(top_widget)
        self.tree.setModel(self.fs_model)
        self.tree.setRootIndex(self.fs_model.index(str(self.pdf_root)))
        self.tree.setAnimated(True)
        self.tree.setSortingEnabled(True)
        header = self.tree.header()
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(80)
        try:
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        except Exception:
            header.setResizeMode(0, QHeaderView.ResizeToContents)
        self.tree.doubleClicked.connect(self._open_from_tree)
        self.tree.selectionModel().selectionChanged.connect(self._tree_selection_changed)
        top_layout.addWidget(self.tree)
    
        # Dole detail
        bottom_widget = QWidget(vsplit)
        bottom_layout = QVBoxLayout(bottom_widget)
    
        scroll = QScrollArea(bottom_widget)
        scroll.setWidgetResizable(True)
        form_host = QWidget(scroll)
        self.detail_form = QFormLayout(form_host)
        self.detail_form.setRowWrapPolicy(QFormLayout.WrapLongRows)
        self.detail_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.detail_form.setFormAlignment(Qt.AlignTop)
        self.detail_form.setLabelAlignment(Qt.AlignRight | Qt.AlignTop)
    
        def _mklabel() -> QLabel:
            lab = QLabel("-")
            lab.setWordWrap(True)
            lab.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
            return lab
    
        # Původní (ponechávám identické názvy)
        self.lbl_board = getattr(self, "lbl_board", _mklabel())
        self.lbl_known = getattr(self, "lbl_known", _mklabel())
        self.lbl_app_type = getattr(self, "lbl_app_type", _mklabel())
        self.lbl_inst = getattr(self, "lbl_inst", _mklabel())
        self.lbl_cand = getattr(self, "lbl_cand", _mklabel())
        self.lbl_acad = getattr(self, "lbl_acad", _mklabel())
        self.lbl_cert = getattr(self, "lbl_cert", _mklabel())
        self.lbl_contact = getattr(self, "lbl_contact", _mklabel())
        self.lbl_email = getattr(self, "lbl_email", _mklabel())
        self.lbl_phone = getattr(self, "lbl_phone", _mklabel())
        self.lbl_postal = getattr(self, "lbl_postal", _mklabel())
        # NOVÉ + původní
        self.lbl_printed_name_title = getattr(self, "lbl_printed_name_title", _mklabel())
        self.lbl_date = getattr(self, "lbl_date", _mklabel())  # Signature Date
        self.lbl_rmb = getattr(self, "lbl_rmb", _mklabel())   # Receiving Member Board
        self.lbl_date_received = getattr(self, "lbl_date_received", _mklabel())
        self.lbl_validity_start = getattr(self, "lbl_validity_start", _mklabel())
        self.lbl_validity_end = getattr(self, "lbl_validity_end", _mklabel())
        self.lbl_syllabi = getattr(self, "lbl_syllabi", _mklabel())
        self.lbl_courses = getattr(self, "lbl_courses", _mklabel())
        self.lbl_proof = getattr(self, "lbl_proof", _mklabel())
        self.lbl_links = getattr(self, "lbl_links", _mklabel())
        self.lbl_additional = getattr(self, "lbl_additional", _mklabel())
        self.lbl_sorted_status = getattr(self, "lbl_sorted_status", _mklabel())
    
        # Sestavení formuláře (pořadí – Printed Name před Signature Date; sekce 7 za ním)
        self.detail_form.addRow("Board:", self.lbl_board)
        self.detail_form.addRow("Application Type:", self.lbl_app_type)
        self.detail_form.addRow("Institution Name:", self.lbl_inst)
        self.detail_form.addRow("Candidate Name:", self.lbl_cand)
        self.detail_form.addRow("Academia Recognition:", self.lbl_acad)
        self.detail_form.addRow("Certified Recognition:", self.lbl_cert)
        self.detail_form.addRow("Full Name:", self.lbl_contact)
        self.detail_form.addRow("Email Address:", self.lbl_email)
        self.detail_form.addRow("Phone Number:", self.lbl_phone)
        self.detail_form.addRow("Postal Address:", self.lbl_postal)
        # Consent
        self.detail_form.addRow("Printed Name, Title:", self.lbl_printed_name_title)
        self.detail_form.addRow("Signature Date:", self.lbl_date)
        # ISTQB internal
        self.detail_form.addRow("Receiving Member Board:", self.lbl_rmb)
        self.detail_form.addRow("Date Received:", self.lbl_date_received)
        self.detail_form.addRow("Validity Start Date:", self.lbl_validity_start)
        self.detail_form.addRow("Validity End Date:", self.lbl_validity_end)
        # Dlouhé texty
        self.detail_form.addRow("Syllabi Integration:", self.lbl_syllabi)
        self.detail_form.addRow("Courses/Modules:", self.lbl_courses)
        self.detail_form.addRow("Proof of ISTQB Certifications:", self.lbl_proof)
        self.detail_form.addRow("University Links:", self.lbl_links)
        self.detail_form.addRow("Additional Info/Documents:", self.lbl_additional)
    
        scroll.setWidget(form_host)
        bottom_layout.addWidget(scroll)
    
        vsplit.addWidget(top_widget)
        vsplit.addWidget(bottom_widget)
        vsplit.setStretchFactor(0, 2)
        vsplit.setStretchFactor(1, 1)
        vsplit.setCollapsible(1, False)
    
        from PySide6.QtWidgets import QVBoxLayout as _VBL
        outer = _VBL(self.browser_tab)
        outer.addWidget(vsplit)

    # ----- Data -----
    def rescan(self) -> None:
        """Scan PDF root and repopulate the Overview table while preserving selection.
        Minimal-change: přidány nové sloupce, „Sorted“ zůstává poslední a je dopočten původními funkcemi.
        """
        from pathlib import Path
        from PySide6.QtCore import Qt, QItemSelectionModel
        from PySide6.QtGui import QStandardItemModel, QStandardItem, QBrush, QColor
        from PySide6.QtWidgets import QStyle
    
        headers = [
            "Board",
            "Application\nApplication Type",
            "Name of Your Academic Institution\nInstitution Name",
            "Name of Your Academic Institution\nCandidate Name",
            "Wished Recognitions\nAcademia Recognition",
            "Wished Recognitions\nCertified Recognition",
            "Contact details for Information exchange\nFull Name",
            "Contact details for Information exchange\nEmail Address",
            "Contact details for Information exchange\nPhone Number",
            "Contact details for Information exchange\nPostal Address",
            "Eligibility Evidence\nSyllabi Integration",
            "Eligibility Evidence\nCourses/Modules",
            "Eligibility Evidence\nProof of ISTQB Certifications",
            "Eligibility Evidence\nUniversity Links",
            "Eligibility Evidence\nAdditional Info/Documents",
            "Declaration and Consent\nPrinted Name, Title",
            "Signature Date",
            "For ISTQB Academia Purpose Only\nReceiving Member Board",
            "For ISTQB Academia Purpose Only\nDate Received",
            "For ISTQB Academia Purpose Only\nValidity Start Date",
            "For ISTQB Academia Purpose Only\nValidity End Date",
            "File\nFile name",
            "Sorted",
        ]
        self._headers = headers
    
        # Pomocná funkce: index podle „tailu“
        def find_col_tail(hdrs: list[str], tail: str) -> int | None:
            for i, h in enumerate(hdrs):
                t = h.split("\n")[-1].strip() if "\n" in h else h.strip()
                if t.lower() == tail.lower():
                    return i
            return None
    
        FILE_COL = find_col_tail(headers, "File name")
        SORTED_COL = find_col_tail(headers, "Sorted")
    
        # Ulož aktuální výběr (podle „File name“ z PROXY)
        selected_paths: set[str] = set()
        try:
            if hasattr(self, "table") and self.table.selectionModel() and self._proxy:
                for pidx in self.table.selectionModel().selectedRows(FILE_COL):
                    key = self._proxy.data(pidx, Qt.UserRole + 1) or self._proxy.data(pidx, Qt.DisplayRole)
                    if key:
                        selected_paths.add(str(key))
        except Exception:
            pass
    
        # Model
        if not hasattr(self, "_source_model"):
            self._source_model = QStandardItemModel(0, len(headers), self)
        else:
            self._source_model.setColumnCount(len(headers))
        self._source_model.setHorizontalHeaderLabels(headers)
        model = self._source_model
        self.table.setModel(self._proxy)  # proxy už je nastaven v _build_overview_tab
    
        # Root
        root = None
        try:
            root = Path(self.pdf_root) if isinstance(self.pdf_root, (str, Path)) else None
            if root and not root.exists():
                root = None
        except Exception:
            root = None
        if not root:
            model.removeRows(0, model.rowCount())
            try:
                self.statusBar().showMessage("PDF root not found.")
            except Exception:
                pass
            return
    
        # Scan
        from app.pdf_scanner import PdfScanner
        scanner = PdfScanner(root)
        self.records = scanner.scan()
    
        # Vyprázdni a naplň
        if model.rowCount() > 0:
            model.removeRows(0, model.rowCount())
    
        # Barevné skupiny (původní + nové)
        COLS_APPLICATION = [1]
        COLS_INSTITUTION = [2, 3]
        COLS_RECOG      = [4, 5]
        COLS_CONTACT    = [6, 7, 8, 9]
        COLS_ELIG       = [10, 11, 12, 13, 14]
        COLS_CONSENT    = [15, 16]          # NOVĚ
        COLS_ISTQB_INT  = [17, 18, 19, 20]  # NOVĚ
    
        BRUSH_APP   = QBrush(QColor(58, 74, 110))
        BRUSH_INST  = QBrush(QColor(74, 58, 110))
        BRUSH_RECOG = QBrush(QColor(58, 110, 82))
        BRUSH_CONT  = QBrush(QColor(110, 82, 58))
        BRUSH_ELIG  = QBrush(QColor(92, 92, 92))
        BRUSH_CONS  = QBrush(QColor(110, 58, 74))
        BRUSH_ISTQB = QBrush(QColor(58, 92, 110))
    
        icon_yes = self.style().standardIcon(QStyle.SP_DialogApplyButton)
        icon_no  = self.style().standardIcon(QStyle.SP_DialogCancelButton)
    
        def paint_group(items: list[QStandardItem], cols: list[int], brush: QBrush) -> None:
            for c in cols:
                if 0 <= c < len(items):
                    items[c].setBackground(brush)
    
        def set_yesno_icon(item: QStandardItem) -> None:
            val = (item.text() or "").strip().lower()
            item.setIcon(icon_yes if val in {"yes", "on", "true", "1", "checked"} else icon_no)
    
        found = 0
        for rec in self.records:
            row_vals = rec.as_row()  # odpovídá headers KROMĚ „Sorted“
            items = [QStandardItem(v) for v in row_vals]
            for it in items:
                it.setEditable(False)
            # doplň prázdný „Sorted“
            items.append(QStandardItem(""))
            paint_group(items, COLS_APPLICATION, BRUSH_APP)
            paint_group(items, COLS_INSTITUTION, BRUSH_INST)
            paint_group(items, COLS_RECOG,      BRUSH_RECOG)
            paint_group(items, COLS_CONTACT,    BRUSH_CONT)
            paint_group(items, COLS_ELIG,       BRUSH_ELIG)
            paint_group(items, COLS_CONSENT,    BRUSH_CONS)
            paint_group(items, COLS_ISTQB_INT,  BRUSH_ISTQB)
            set_yesno_icon(items[4])  # Academia
            set_yesno_icon(items[5])  # Certified
    
            # ulož absolutní cestu do „File name“ (UserRole+1)
            if FILE_COL is not None and 0 <= FILE_COL < len(items):
                items[FILE_COL].setData(str(rec.path), Qt.UserRole + 1)
    
            model.appendRow(items)
            found += 1
    
        # třídění a fit
        self.table.sortByColumn(0, Qt.AscendingOrder)
        try:
            for c in range(len(headers)):
                self.table.resizeColumnToContents(c)
            for c in (10, 11, 12, 13, 14):
                self.table.setColumnHidden(c, True)
        except Exception:
            pass
    
        # označ „Sorted“ a aplikuj hiding (původní logika)
        try:
            self._overview_update_sorted_flags()
            self._overview_apply_sorted_row_hiding()
        except Exception:
            pass
    
        # obnov výběr
        try:
            if selected_paths and self._proxy is not None:
                sel_model = self.table.selectionModel()
                for r in range(model.rowCount()):
                    sval = model.index(r, FILE_COL).data(Qt.UserRole + 1) if FILE_COL is not None else None
                    if not sval and FILE_COL is not None:
                        sval = model.index(r, FILE_COL).data()
                    if sval and str(sval) in selected_paths:
                        pidx = self._proxy.mapFromSource(model.index(r, 0))
                        sel_model.select(pidx, QItemSelectionModel.Select | QItemSelectionModel.Rows)
        except Exception:
            pass
    
        self._rebuild_watch_list()
        try:
            root_str = str(self.pdf_root) if isinstance(self.pdf_root, Path) else "<unset>"
            self.statusBar().showMessage(f"PDF found: {found} • Parsed: {len(self.records)} • Root: {root_str}")
        except Exception:
            pass

    # ----- Actions -----
    from typing import Optional
    from PySide6.QtCore import Qt, QModelIndex
    from PySide6.QtWidgets import QTableView
    from PySide6.QtCore import QSortFilterProxyModel
    
    def _selected_record(self) -> Optional[PdfRecord]:
        """
        Return PdfRecord for the *Overview* table selection.
        Robust to multiple table views in other tabs:
        - If invoked from a view signal, use sender() if it's a QTableView.
        - Else, find the Overview table inside self.overview_tab.
        """
        # Prefer the signal sender if it's the table view
        view = None
        snd = self.sender()
        if isinstance(snd, QTableView):
            view = snd
        else:
            # Fallback: locate the Overview table within the Overview tab
            try:
                if hasattr(self, "overview_tab") and self.overview_tab is not None:
                    view = self.overview_tab.findChild(QTableView)  # no objectName required
            except Exception:
                view = None
            # Last resort: use self.table if it is a QTableView
            if view is None and isinstance(getattr(self, "table", None), QTableView):
                view = self.table
    
        if view is None:
            return None
    
        sel = view.selectionModel()
        if not sel or not sel.hasSelection():
            return None
    
        # First selected row in the proxy model
        pindex = sel.selectedRows()[0]
        proxy = view.model()
        if isinstance(proxy, QSortFilterProxyModel):
            # Map selected proxy row to source row; column doesn't matter for row mapping
            srow = proxy.mapToSource(proxy.index(pindex.row(), 0)).row()
            src = proxy.sourceModel()
            file_col = src.columnCount() - 1  # last column stores filename/full path
            idx = src.index(srow, file_col)
            path_str = idx.data(Qt.UserRole + 1) or idx.data()
        else:
            src = proxy
            file_col = src.columnCount() - 1
            idx = src.index(pindex.row(), file_col)
            path_str = idx.data(Qt.UserRole + 1) or idx.data()
    
        if not path_str:
            return None
    
        # Match against loaded records
        for r in getattr(self, "records", []):
            if str(r.path) == str(path_str):
                return r
        return None
    
    def _init_fs_watcher(self) -> None:
        from PySide6.QtCore import QTimer
        from PySide6.QtCore import QFileSystemWatcher

        self._watcher = QFileSystemWatcher(self)
        self._watcher.directoryChanged.connect(self._on_fs_changed)
        self._watcher.fileChanged.connect(self._on_fs_changed)

        self._fs_debounce = QTimer(self)
        self._fs_debounce.setSingleShot(True)
        self._fs_debounce.setInterval(300)
        self._fs_debounce.timeout.connect(self._fs_debounced)

        self._rebuild_watch_list()

    def _rebuild_watch_list(self) -> None:
        """Watch PDF root, all subdirs, and all .pdf files."""
        if not hasattr(self, "_watcher"):
            return
        watcher = self._watcher
        # Clear
        try:
            if watcher.files():
                watcher.removePaths(watcher.files())
            if watcher.directories():
                watcher.removePaths(watcher.directories())
        except Exception:
            pass

        dirs = set()
        files = set()
        if self.pdf_root.exists():
            dirs.add(str(self.pdf_root))
            for p in self.pdf_root.rglob("*"):
                if p.is_dir():
                    dirs.add(str(p))
                elif p.suffix.lower() == ".pdf":
                    files.add(str(p))
        if dirs:
            watcher.addPaths(sorted(dirs))
        if files:
            watcher.addPaths(sorted(files))

    def _on_fs_changed(self, _path: str) -> None:
        # Debounce frequent events
        if hasattr(self, "_fs_debounce"):
            self._fs_debounce.start()

    def _fs_debounced(self) -> None:
        # Rescan and refresh watchers (new subdirs/files)
        self.rescan()
        self._rebuild_watch_list()


    from typing import Optional
    from PySide6.QtWidgets import QMessageBox
    from PySide6.QtGui import QDesktopServices
    from PySide6.QtCore import QUrl   
    
    def open_selected_pdf(self) -> None:
        rec = self._selected_record()
        if not rec:
            QMessageBox.information(self, "Open PDF", "Please select a row first.")
            return
        # 0.6d: QDesktopServices.openUrl vyžaduje QUrl; posílejme lokální souborový URL
        QDesktopServices.openUrl(self.QUrl.fromLocalFile(str(rec.path)))

    def _open_from_tree(self, index) -> None:
        path = Path(self.fs_model.filePath(index))
        if path.is_file() and path.suffix.lower() == ".pdf":
            QDesktopServices.openUrl(path.as_uri())

    def _open_selected_detail(self) -> None:
        # open currently selected file in the tree
        sel = self.tree.selectionModel()
        if not sel or not sel.hasSelection():
            QMessageBox.information(self, "Open PDF", "Please select a file in the tree.")
            return
        idx = sel.selectedRows()[0]
        self._open_from_tree(idx)

    def _tree_selection_changed(self, *_):
        sel = self.tree.selectionModel()
        if not sel or not sel.hasSelection():
            return
        idx = sel.selectedRows()[0]
        path = Path(self.fs_model.filePath(idx))
        if not path.is_file():
            return

        from .pdf_scanner import PdfScanner
        scanner = PdfScanner(self.pdf_root)
        rec: Optional[PdfRecord] = None
        try:
            rec = scanner._parse_one(path)
        except Exception:
            rec = None

        self._update_detail_panel(rec)

    def _renumber_rows(self) -> None:
        """Write 1..N into the 'No.' column in source model according to current proxy order."""
        proxy = self.table.model()
        if not isinstance(proxy, QSortFilterProxyModel):
            return
        src = proxy.sourceModel()
        if src is None:
            return
        # NUM col = 0
        for r in range(proxy.rowCount()):
            pidx = proxy.index(r, 0)  # any column maps row; col 0 exists
            sidx = proxy.mapToSource(pidx)
            # Set display text to 1-based row number
            src.setData(src.index(sidx.row(), 0), str(r + 1), Qt.DisplayRole)

    def _update_detail_panel(self, rec: Optional[PdfRecord]) -> None:
        if rec is None:
            for lbl in (
                self.lbl_board, self.lbl_known, self.lbl_app_type, self.lbl_inst, self.lbl_cand,
                self.lbl_acad, self.lbl_cert, self.lbl_contact, self.lbl_email, self.lbl_phone,
                self.lbl_postal, self.lbl_printed_name_title, self.lbl_date, self.lbl_rmb,
                self.lbl_date_received, self.lbl_validity_start, self.lbl_validity_end,
                self.lbl_syllabi, self.lbl_courses, self.lbl_proof, self.lbl_links, self.lbl_additional
            ):
                lbl.setText("-")
            return
    
        self.lbl_board.setText(rec.board)
        self.lbl_known.setText("Yes" if rec.board_known else "Unverified")
        self.lbl_app_type.setText(rec.application_type or "")
        self.lbl_inst.setText(rec.institution_name or "")
        self.lbl_cand.setText(rec.candidate_name or "")
        self.lbl_acad.setText(rec.recognition_academia or "")
        self.lbl_cert.setText(rec.recognition_certified or "")
        self.lbl_contact.setText(rec.contact_full_name or "")
        self.lbl_email.setText(rec.contact_email or "")
        self.lbl_phone.setText(rec.contact_phone or "")
        self.lbl_postal.setText(rec.contact_postal_address or "")
        # NOVÉ
        self.lbl_printed_name_title.setText(rec.printed_name_title or "")
        self.lbl_date.setText(rec.signature_date or "")
        self.lbl_rmb.setText(rec.receiving_member_board or "")
        self.lbl_date_received.setText(rec.date_received or "")
        self.lbl_validity_start.setText(rec.validity_start_date or "")
        self.lbl_validity_end.setText(rec.validity_end_date or "")
        # Dlouhé texty
        self.lbl_syllabi.setText(rec.syllabi_integration_description or "")
        self.lbl_courses.setText(rec.courses_modules_list or "")
        self.lbl_proof.setText(rec.proof_of_istqb_certifications or "")
        self.lbl_links.setText(rec.university_links or "")
        self.lbl_additional.setText(rec.additional_information_documents or "")
        
        
        
        
        
        
    def showEvent(self, event) -> None:
        """
        Jednorázové globální sizing hinty při prvním zobrazení okna.
        Qt volá showEvent automaticky. Zvětší okno alespoň o 50 % na šířku i výšku,
        ale s respektem k dostupnému pracovního prostoru obrazovky.
        """
        try:
            super().showEvent(event)
        except Exception:
            try:
                super(type(self), self).showEvent(event)
            except Exception:
                pass
    
        if getattr(self, "_sizing_applied", False):
            return
    
        try:
            self._apply_global_sizing_once()
        finally:
            self._sizing_applied = True
            
    def _apply_global_sizing_once(self) -> None:
        """
        Orchestrátor fit-to-data:
          1) zvětšení hlavního okna min. o 50 % (s limitem dle obrazovky),
          2) lokální fit v Overview,
          3) lokální fit ve Sorted PDFs,
          4) lokální fit v PDF Browseru (pokud existuje helper),
          5) zopakování po event loopu.
        """
        from PySide6.QtGui import QGuiApplication
        from PySide6.QtCore import QTimer
    
        # 1) Zvětšení okna min. o 50 %; respektovat dostupný prostor obrazovky
        try:
            cw, ch = max(self.width(), 800), max(self.height(), 600)
            desired_w = int(cw * 1.5)
            desired_h = int(ch * 1.5)
    
            scr = QGuiApplication.primaryScreen()
            if scr:
                avail = scr.availableGeometry()
                max_w = int(avail.width() * 0.95)
                max_h = int(avail.height() * 0.95)
                desired_w = min(desired_w, max_w)
                desired_h = min(desired_h, max_h)
    
            # Nepoužij menší než aktuální
            desired_w = max(desired_w, cw)
            desired_h = max(desired_h, ch)
            self.resize(desired_w, desired_h)
        except Exception:
            pass
    
        # 2–4) Lokální fit helpery
        for fn in (getattr(self, "_apply_overview_sizes", None),
                   getattr(self, "_apply_sorted_sizes", None),
                   getattr(self, "_apply_browser_sizes", None)):
            if callable(fn):
                try:
                    fn()
                except Exception:
                    pass
    
        # 5) Po event loopu zopakuj (kvůli opožděnému plnění modelů)
        def _post():
            for fn in (getattr(self, "_apply_overview_sizes", None),
                       getattr(self, "_apply_sorted_sizes", None),
                       getattr(self, "_apply_browser_sizes", None)):
                if callable(fn):
                    try:
                        fn()
                    except Exception:
                        pass
        QTimer.singleShot(0, _post)
        
    def _apply_overview_sizes(self) -> None:
        """
        Fit tabulky v Overview k aktuálním datům.
        Nemění výběry, sort ani filtry; jen rozměry sloupců a roztažení posledního.
        """
        from PySide6.QtCore import QTimer
        view = getattr(self, "table", None)
        if not view:
            return
        try:
            hh = view.horizontalHeader()
            hh.setStretchLastSection(True)
            view.resizeColumnsToContents()
            QTimer.singleShot(0, view.resizeColumnsToContents)
        except Exception:
            pass
        
    def _apply_browser_sizes(self) -> None:
        """
        Fit v "PDF Browser": sloupce stromu dle obsahu, včetně re-fit po načtení adresářů.
        Zachovává existující chování (_build_browser_tab už řeší Name/ResizeToContents).
        Tady navíc:
          - resize všech dostupných sloupců podle obsahu,
          - jednorázové napojení na fs_model.directoryLoaded -> re-fit,
          - žádná změna layoutu/detail panelu.
        """
        from PySide6.QtWidgets import QHeaderView
        from PySide6.QtCore import QTimer
    
        tree = getattr(self, "tree", None)
        fs_model = getattr(self, "fs_model", None)
        if not tree or not fs_model:
            return
    
        header = tree.header()
        try:
            cols = header.count()
        except Exception:
            cols = 1
    
        # Nastav ResizeToContents na všech sloupcích, a hned je přepočti
        for c in range(cols):
            try:
                header.setSectionResizeMode(c, QHeaderView.ResizeToContents)
            except Exception:
                try:
                    header.setResizeMode(c, QHeaderView.ResizeToContents)  # Qt5 fallback
                except Exception:
                    pass
            try:
                tree.resizeColumnToContents(c)
            except Exception:
                pass
    
        # Druhé kolo po event loopu
        def _second_pass():
            for c in range(cols):
                try:
                    tree.resizeColumnToContents(c)
                except Exception:
                    pass
        QTimer.singleShot(0, _second_pass)
    
        # Napojení na directoryLoaded (jen jednou)
        if not getattr(self, "_browser_sizes_connected", False):
            try:
                def _on_loaded(*_):
                    QTimer.singleShot(0, self._apply_browser_sizes)
                fs_model.directoryLoaded.connect(_on_loaded)
                self._browser_sizes_connected = True
            except Exception:
                pass
        
    def _apply_sorted_sizes(self) -> None:
        """
        Fit ve "Sorted PDFs": zvětšený levý panel (Board / PDF) a sloupec,
        jemný poměr splitteru ~65 % : 35 %, autosize sloupce a malé navýšení šířky.
        """
        from PySide6.QtCore import QTimer
    
        tree = getattr(self, "tree_sorted", None)
        splitter = getattr(self, "split_sorted", None)
    
        if tree:
            try:
                tree.resizeColumnToContents(0)
                # po prvním autosize lehce přidej rezervu
                extra = 80
                curr = tree.columnWidth(0)
                if curr > 0:
                    tree.setColumnWidth(0, curr + extra)
                QTimer.singleShot(0, lambda: tree.resizeColumnToContents(0))
            except Exception:
                pass
    
        if splitter:
            try:
                total_w = max(getattr(self, "width", lambda: 1400)(), 1400)
                # ~65 % vlevo (strom), ~35 % vpravo (formulář)
                splitter.setSizes([int(total_w * 0.65), int(total_w * 0.35)])
            except Exception:
                pass
            
